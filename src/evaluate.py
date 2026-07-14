"""
src/evaluate.py
===============
Compare CNN-only vs rule-engine predictions against ground truth.

Saves: results/evaluation_results.xlsx

KEY DESIGN DECISIONS
────────────────────
Tier semantics (must match pyreason_engine.py):
  Tier 1  Confirmed  : HIGH-conf, firm, not changed, not flagged → spot-check
  Tier 2  Suggested  : Rule 1/2 corrected it (changed=True), OR
                       Rule 1/2 fired and spatial evidence supports the
                       existing label (changed=False but rule1/rule2 in rule_applied)
                       → suggested label available; confirm or override
  Tier 3  Attention  : Rule 3 cluster flag
                       → spatially coherent uncertainty/minority zone; review collectively
  Tier 4  Refer      : Rule 4 isolated LOW/Mixed flag
                       → full human decision; no separate Rule 4/4b split

Rule-tier mapping:
  Rule 1a/1b corrected patch  → Tier 2  (changed=True, needs_review=False)
  Rule 2  corrected patch     → Tier 2  (changed=True, needs_review=False)
  Rule 3  cluster flagged     → Tier 3  (changed=False, needs_review=True)
  Rule 4  refer flagged       → Tier 4  (changed=False, needs_review=True)
  Uncorrected HIGH firm       → Tier 1
  Uncorrected MEDIUM/LOW firm → Tier 4  (no spatial rule acted)

Why Rule 4 does NOT stack onto spatial corrections or Rule 3:
  If Rule 1/2 already corrected a LOW/Mixed patch using strong spatial evidence,
  adding needs_review=True contradicts the correction and inflates referral counts.
  If Rule 3 already flagged a patch, Rule 4 is suppressed so the cluster context is preserved.
  Corrected patches are tracked via correction_confidence tag instead.

Rule 3 candidates:
  Rule 3 can include Mixed/deferred predictions, LOW-confidence firm predictions,
  and firm minority predictions of any confidence, provided they form a spatial cluster
  and were not already corrected by Rule 1/2.

3-zone prediction evaluation:
  Zone 1  FIRM:     ML or PR output is G or Gplus → binary metrics (MCC, BalAcc, F1…)
  Zone 2  DEFERRED: ML or PR output is Mixed on G/Gplus GT → deferral rate, coverage %
  Zone 3  Mixed GT: Mixed GT patches (training eval; often empty at test)

Primary workload metric: error density per tier (Tier 3+4 >> Tier 1 = triage working).

Excel sheets (in order):
  Summary          — Table B (deferral) + confusion matrix
                     + rule breakdown + correction quality + tier workload
  Workload         — per-tier error density, NNR, enrichment, triage quality rating
  Coverage         — image type breakdown, rule firing counts, silent failures
  Triage Tiers     — tier summary + rule breakdown by tier
  Tier 1–4 detail  — one sheet per tier, full patch detail
  Patch Results    — every patch with all columns including triage_tier
  Changed Patches  — only corrected patches (Tier 2 spatial corrections)
  Review Flags     — Rule 3 (Tier 3) and Rule 4 (Tier 4) flagged patches
  Errors Remaining — firm wrong patches not flagged (silent failures)
  Workload & Triage
"""

from __future__ import annotations

import os
from typing import Dict, List, Optional

import numpy as np
import pandas as pd

import config


# ─────────────────────────────────────────────────────────────────
#  Constants
# ─────────────────────────────────────────────────────────────────

BINARY_LABELS    = {"G", "Gplus"}
UNCERTAIN_LABELS = {"Mixed", "Mix", "mix", "mixed"}


# ─────────────────────────────────────────────────────────────────
#  Slide final  (with REFER support)
# ─────────────────────────────────────────────────────────────────

def compute_slide_final(slides: Dict, final: Dict, slide_vote: Dict) -> Dict:
    """
    Compute per-slide diagnosis after spatial rule engine.

    diagnosis values:
      "Gplus"  — majority of final predictions are Gplus
      "G"      — majority of final predictions are G
      "REFER"  — majority of final predictions are Mixed (uncertain → refer)
      "TIE"    — equal Gplus and G, no Mixed majority
    """
    slide_final: Dict = {}
    for sid, names in slides.items():
        fps = [final[n] for n in names if n in final]
        if not fps:
            continue
        n_g  = sum(1 for p in fps if p["final_predicted"] == "G")
        n_gp = sum(1 for p in fps if p["final_predicted"] == "Gplus")
        n_mx = sum(1 for p in fps if p["final_predicted"] in UNCERTAIN_LABELS)
        n_ch = sum(1 for p in fps if p["changed"])
        n_rv = sum(1 for p in fps if p["needs_review"])

        if n_mx > n_g and n_mx > n_gp:
            diagnosis = "REFER"
        elif n_gp > n_g:
            diagnosis = "Gplus"
        elif n_g > n_gp:
            diagnosis = "G"
        else:
            diagnosis = "TIE"

        gt_vals  = [final[n]["gt_label"] for n in names
                    if n in final and final[n].get("gt_label") is not None]
        slide_gt = max(set(gt_vals), key=gt_vals.count) if gt_vals else None

        slide_final[sid] = dict(
            diagnosis=diagnosis, slide_gt=slide_gt,
            n_G=n_g, n_Gplus=n_gp, n_Mixed=n_mx, total=len(fps),
            n_changed=n_ch, n_needs_review=n_rv,
            ratio=slide_vote[sid]["ratio"],
            dominant=slide_vote[sid]["dominant"],
        )
    return slide_final


# ─────────────────────────────────────────────────────────────────
#  Metrics helpers
# ─────────────────────────────────────────────────────────────────

def _label_to_int(label: Optional[str]) -> Optional[int]:
    if label is None:
        return None
    return config.LABEL_MAP_INV.get(label)


def _cls_metrics(y_true: List[int], y_pred: List[int]) -> Dict:
    """
    Binary classification metrics (G=0 vs Gplus=1).
    Only call this on FIRM predictions — i.e. GT in {G, Gplus} AND
    predicted in {G, Gplus}.  Do NOT pass deferred (Mixed) predictions here.
    """
    if not y_true:
        return dict(acc=0, precision=0, recall=0, f1=0, f2=0,
                    mcc=0, bal_acc=0, tp=0, tn=0, fp=0, fn=0, n=0)
    yt  = np.array(y_true)
    yp  = np.array(y_pred)
    n   = len(yt)
    acc = float((yt == yp).sum() / n)
    tp  = int(((yp == 1) & (yt == 1)).sum())
    tn  = int(((yp == 0) & (yt == 0)).sum())
    fp  = int(((yp == 1) & (yt == 0)).sum())
    fn  = int(((yp == 0) & (yt == 1)).sum())
    prec    = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    rec     = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1      = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0
    f2      = (5 * prec * rec) / (4 * prec + rec) if (4 * prec + rec) > 0 else 0.0
    denom   = ((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn)) ** 0.5
    mcc     = (tp*tn - fp*fn) / denom if denom > 0 else 0.0
    spec    = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    bal_acc = (rec + spec) / 2
    return dict(acc=acc, precision=prec, recall=rec, f1=f1, f2=f2,
                mcc=mcc, bal_acc=bal_acc,
                tp=tp, tn=tn, fp=fp, fn=fn, n=n)


def _compute_deferral_stats(
    rows_binary: List[Dict],
    pred_key_ml: str = "ml_predicted",
    pred_key_pr: str = "final_predicted",
) -> Dict:
    n = len(rows_binary)
    if n == 0:
        return dict(n_total=0, n_ml_firm=0, n_ml_deferred=0,
                    ml_deferral_rate=0.0, ml_coverage=0.0,
                    n_pr_firm=0, n_pr_deferred=0,
                    pr_deferral_rate=0.0, pr_coverage=0.0)

    n_ml_def  = sum(1 for r in rows_binary if r[pred_key_ml] in UNCERTAIN_LABELS)
    n_pr_def  = sum(1 for r in rows_binary if r[pred_key_pr] in UNCERTAIN_LABELS)
    n_ml_firm = n - n_ml_def
    n_pr_firm = n - n_pr_def

    return dict(
        n_total=n,
        n_ml_firm=n_ml_firm,
        n_ml_deferred=n_ml_def,
        ml_deferral_rate=n_ml_def / n,
        ml_coverage=n_ml_firm / n,
        n_pr_firm=n_pr_firm,
        n_pr_deferred=n_pr_def,
        pr_deferral_rate=n_pr_def / n,
        pr_coverage=n_pr_firm / n,
    )


# ─────────────────────────────────────────────────────────────────
#  Fair comparison helper
# ─────────────────────────────────────────────────────────────────

def _compute_fair_comparisons(rows: List[Dict]) -> Dict:
    BINARY = BINARY_LABELS
    LABEL  = {"G": 0, "Gplus": 1}

    def _metrics_from_pairs(pairs):
        if not pairs:
            return _cls_metrics([], [])
        yt = [p[0] for p in pairs]
        yp = [p[1] for p in pairs]
        return _cls_metrics(yt, yp)

    binary_rows = [r for r in rows if r.get("gt_label") in BINARY]
    if not binary_rows:
        return {}

    ml_firm_names = {r["image_name"] for r in binary_rows
                     if r.get("ml_predicted") in BINARY}

    # ── Scenario A ─────────────────────────────────────────────────
    ml_firm_rows    = [r for r in binary_rows if r["image_name"] in ml_firm_names]
    ml_a_pairs      = [(LABEL[r["gt_label"]], LABEL[r["ml_predicted"]])
                       for r in ml_firm_rows if r["ml_predicted"] in BINARY]
    pr_a_firm       = [r for r in ml_firm_rows if r.get("final_predicted") in BINARY]
    pr_a_pairs      = [(LABEL[r["gt_label"]], LABEL[r["final_predicted"]])
                       for r in pr_a_firm if r["final_predicted"] in BINARY]
    n_pr_a_deferred = len(ml_firm_rows) - len(pr_a_firm)
    m_ml_a = _metrics_from_pairs(ml_a_pairs)
    m_pr_a = _metrics_from_pairs(pr_a_pairs)

    # ── Scenario B ─────────────────────────────────────────────────
    ml_b_pairs = []
    pr_b_pairs = []
    for r in binary_rows:
        gt = LABEL.get(r.get("gt_label"))
        if gt is None:
            continue
        ml = r.get("ml_predicted")
        pr = r.get("final_predicted")
        ml_b_pairs.append((gt, LABEL[ml] if ml in BINARY else 1 - gt))
        pr_b_pairs.append((gt, LABEL[pr] if pr in BINARY else 1 - gt))
    m_ml_b = _metrics_from_pairs(ml_b_pairs)
    m_pr_b = _metrics_from_pairs(pr_b_pairs)

    # ── Scenario C ─────────────────────────────────────────────────
    extra_rows     = [r for r in binary_rows
                      if r["image_name"] not in ml_firm_names
                      and r.get("final_predicted") in BINARY]
    extra_correct  = sum(1 for r in extra_rows if r.get("pr_correct") is True)
    extra_wrong    = len(extra_rows) - extra_correct
    extra_accuracy = extra_correct / len(extra_rows) if extra_rows else 0.0
    extra_pairs    = [(LABEL[r["gt_label"]], LABEL[r["final_predicted"]])
                      for r in extra_rows if r["final_predicted"] in BINARY]
    m_extra = _metrics_from_pairs(extra_pairs)

    return dict(
        scenario_A=dict(
            description="Same N as ML — apples-to-apples on ML-firm patches",
            n=len(ml_a_pairs),
            n_pr_deferred_on_ml_set=n_pr_a_deferred,
            ml=m_ml_a, pr=m_pr_a,
            delta_mcc=m_pr_a.get("mcc", 0) - m_ml_a.get("mcc", 0),
            delta_recall=m_pr_a.get("recall", 0) - m_ml_a.get("recall", 0),
        ),
        scenario_B=dict(
            description="All patches, Mixed=wrong — full population forced answer",
            n=len(ml_b_pairs),
            ml=m_ml_b, pr=m_pr_b,
            delta_mcc=m_pr_b.get("mcc", 0) - m_ml_b.get("mcc", 0),
            delta_recall=m_pr_b.get("recall", 0) - m_ml_b.get("recall", 0),
        ),
        scenario_C=dict(
            description="Extra patches PR committed (ML deferred, PR firm)",
            n=len(extra_rows),
            n_correct=extra_correct,
            n_wrong=extra_wrong,
            accuracy=extra_accuracy,
            metrics=m_extra,
        ),
    )


# ─────────────────────────────────────────────────────────────────
#  Workload / triage quality metrics
# ─────────────────────────────────────────────────────────────────

def _compute_workload_metrics(
    patch_df: "pd.DataFrame",
    has_gt: bool,
    n_r3_flagged: int,
    n_r3_errors: int,
) -> Dict:
    if not has_gt or len(patch_df) == 0:
        return {}

    n_total = len(patch_df)
    t1  = patch_df[patch_df["triage_tier"] == 1]
    t2  = patch_df[patch_df["triage_tier"] == 2]
    t3  = patch_df[patch_df["triage_tier"] == 3]
    t4  = patch_df[patch_df["triage_tier"] == 4]
    t34 = patch_df[patch_df["triage_tier"].isin([3, 4])]

    n_t1  = len(t1);  n_t2  = len(t2)
    n_t3  = len(t3);  n_t4  = len(t4);  n_t34 = len(t34)

    def _is_cnn_error(df):
        has_gt_mask = df["gt_label"].notna()
        wrong_firm  = (df["ml_correct"] == False) & (df["ml_zone"] == "firm")
        deferred    = df["ml_zone"] == "deferred"
        return df[has_gt_mask & (wrong_firm | deferred)]

    def _is_pr_error(df):
        has_gt_mask = df["gt_label"].notna()
        wrong_firm  = (df["pr_correct"] == False) & (df["pr_zone"] == "firm")
        deferred    = df["pr_zone"] == "deferred"
        return df[has_gt_mask & (wrong_firm | deferred)]

    def _is_handled(df):
        corrected = (df["outcome"] == "improved") & df["gt_label"].notna()
        flagged   = df["needs_review"] & df["gt_label"].notna()
        return df[corrected | flagged]

    def _is_remaining_error(df):
        has_gt_mask  = df["gt_label"].notna()
        is_error     = (
            ((df["pr_correct"] == False) & (df["pr_zone"] == "firm")) |
            (df["pr_zone"] == "deferred")
        )
        not_flagged  = ~df["needs_review"]
        not_improved = df["outcome"] != "improved"
        return df[has_gt_mask & is_error & not_flagged & not_improved]

    all_cnn_errors_df = _is_cnn_error(patch_df)
    all_pr_errors_df  = _is_pr_error(patch_df)
    all_remaining_df  = _is_remaining_error(patch_df)

    n_cnn_errors_total = len(all_cnn_errors_df)
    n_pr_errors_total  = len(all_pr_errors_df)

    n_ml_deferred   = int((patch_df["ml_zone"] == "deferred").sum())
    n_ml_wrong_firm = int(
        ((patch_df["ml_correct"] == False) & (patch_df["ml_zone"] == "firm") &
         patch_df["gt_label"].notna()).sum()
    )

    def _tier_stats(df_tier):
        cnn_err = _is_cnn_error(df_tier)
        pr_err  = _is_pr_error(df_tier)
        handled = _is_handled(df_tier)
        remain  = _is_remaining_error(df_tier)
        n       = len(df_tier)
        n_cnn   = len(cnn_err)
        n_pr    = len(pr_err)
        n_hand  = min(len(handled), n_cnn)
        n_rem   = len(remain)
        density_rate = n_pr / n * 100 if n > 0 else 0.0
        silent_rate  = n_rem / n * 100 if n > 0 else 0.0
        return dict(
            n=n, n_cnn_errors=n_cnn, n_pr_errors=n_pr,
            n_handled=n_hand, n_remaining=n_rem,
            error_rate=round(density_rate, 1),
            density_rate=round(density_rate, 1),
            silent_failure_rate=round(silent_rate, 1),
        )

    t1_stats  = _tier_stats(t1)
    t2_stats  = _tier_stats(t2)
    t3_stats  = _tier_stats(t3)
    t4_stats  = _tier_stats(t4)
    t34_stats = _tier_stats(t34)

    t2_corrected = int(((t2["outcome"] == "improved") & t2["gt_label"].notna()).sum())
    t2_flagged   = int((t2["needs_review"] & t2["gt_label"].notna() & (t2["outcome"] != "improved")).sum())
    t3_flagged   = int((t3["needs_review"] & t3["gt_label"].notna()).sum())
    t4_flagged   = int((t4["needs_review"] & t4["gt_label"].notna()).sum())

    firm_df         = patch_df[(patch_df["pr_zone"] == "firm") & patch_df["gt_label"].notna()]
    all_errors_firm = int((firm_df["pr_correct"] == False).sum())

    t1_firm  = t1[(t1["pr_zone"] == "firm") & t1["gt_label"].notna()]
    t34_firm = t34[(t34["pr_zone"] == "firm") & t34["gt_label"].notna()]
    t2_firm  = t2[(t2["pr_zone"] == "firm") & t2["gt_label"].notna()]
    t3_firm  = t3[(t3["pr_zone"] == "firm") & t3["gt_label"].notna()]
    t4_firm  = t4[(t4["pr_zone"] == "firm") & t4["gt_label"].notna()]

    n_t1_firm  = len(t1_firm)
    n_t34_firm = len(t34_firm)

    t1_errors_firm  = int((t1_firm["pr_correct"]  == False).sum())
    t34_errors_firm = int((t34_firm["pr_correct"] == False).sum())
    t2_errors_firm  = int((t2_firm["pr_correct"]  == False).sum())
    t3_errors_firm  = int((t3_firm["pr_correct"]  == False).sum())
    t4_errors_firm  = int((t4_firm["pr_correct"]  == False).sum())

    pr_deferred_t1  = t1[(t1["pr_zone"] == "deferred") & t1["gt_label"].notna()]
    pr_deferred_t34 = t34[(t34["pr_zone"] == "deferred") & t34["gt_label"].notna()]
    pr_deferred_all = patch_df[(patch_df["pr_zone"] == "deferred") & patch_df["gt_label"].notna()]
    n_pr_deferred_t1  = len(pr_deferred_t1)
    n_pr_deferred_t34 = len(pr_deferred_t34)
    n_pr_deferred_all = len(pr_deferred_all)

    baseline_density  = n_cnn_errors_total / n_total if n_total > 0 else 0.0
    t1_error_density  = t1_stats["n_pr_errors"]  / n_t1  if n_t1  > 0 else 0.0
    t34_error_density = t34_stats["n_pr_errors"] / n_t34 if n_t34 > 0 else 0.0
    t2_error_density  = t2_stats["n_pr_errors"]  / n_t2  if n_t2  > 0 else 0.0
    t3_error_density  = t3_stats["n_pr_errors"]  / n_t3  if n_t3  > 0 else 0.0
    t4_error_density  = t4_stats["n_pr_errors"]  / n_t4  if n_t4  > 0 else 0.0

    if t1_error_density > 0:
        density_ratio     = t34_error_density / t1_error_density
        density_ratio_inf = False
    else:
        density_ratio     = None
        density_ratio_inf = True

    if baseline_density > 0:
        t34_enrichment     = t34_error_density / baseline_density
        t34_enrichment_inf = False
    else:
        t34_enrichment     = None
        t34_enrichment_inf = True

    DENSITY_RATIO_STRONG   = 3.0
    DENSITY_RATIO_ADEQUATE = 1.5
    dr = density_ratio if density_ratio is not None else float("inf")
    triage_quality = (
        "strong"   if dr >= DENSITY_RATIO_STRONG   else
        "adequate" if dr >= DENSITY_RATIO_ADEQUATE else
        "weak"     if dr > 1.0                     else
        "failed"
    )

    spot_check_pct  = n_t1  / n_total * 100 if n_total > 0 else 0.0
    deep_review_pct = n_t34 / n_total * 100 if n_total > 0 else 0.0

    nnr_ml_problems = n_cnn_errors_total
    nnr_ml = n_total / nnr_ml_problems if nnr_ml_problems > 0 else float("inf")

    nnr_pr_problems = t34_stats["n_pr_errors"]
    nnr_pr = n_t34 / nnr_pr_problems if nnr_pr_problems > 0 else float("inf")

    if nnr_pr > 0 and nnr_pr != float("inf") and nnr_ml != float("inf"):
        nnr_improvement = round(nnr_ml / nnr_pr, 2)
    else:
        nnr_improvement = None

    problem_concentration_t34 = (
        t34_stats["n_pr_errors"] / n_pr_errors_total * 100
        if n_pr_errors_total > 0 else 0.0
    )
    error_in_t1_pct = (
        t1_stats["n_pr_errors"] / n_pr_errors_total * 100
        if n_pr_errors_total > 0 else 0.0
    )
    firm_conc_t34 = (
        t34_errors_firm / all_errors_firm * 100
        if all_errors_firm > 0 else 0.0
    )
    firm_t1_pct = (
        t1_errors_firm / all_errors_firm * 100
        if all_errors_firm > 0 else 0.0
    )

    flagged_df         = patch_df[patch_df["needs_review"] & patch_df["gt_label"].notna()]
    n_flagged_total    = len(flagged_df)
    n_flagged_errors   = len(_is_cnn_error(flagged_df))
    flag_precision_all = n_flagged_errors / n_flagged_total * 100 if n_flagged_total > 0 else 0.0

    r3_true_error_rate  = n_r3_errors / n_r3_flagged * 100 if n_r3_flagged > 0 else 0.0
    r3_false_alarm_rate = 100.0 - r3_true_error_rate

    flag_only_df       = patch_df[patch_df["needs_review"] & ~patch_df["changed"] & patch_df["gt_label"].notna()]
    n_flag_only        = len(flag_only_df)
    n_flag_only_errors = len(_is_cnn_error(flag_only_df))

    t2_changed_df = t2[t2["changed"] & t2["gt_label"].notna()]
    n_t2_improved = int((t2_changed_df["outcome"] == "improved").sum())
    n_t2_worsened = int((t2_changed_df["outcome"] == "error").sum())
    n_t2_changed  = n_t2_improved + n_t2_worsened
    t2_correction_precision = (
        round(n_t2_improved / n_t2_changed * 100, 1) if n_t2_changed > 0 else 0.0
    )

    return dict(
        n_total=n_total,
        n_tier1=n_t1, n_tier2=n_t2, n_tier3=n_t3, n_tier4=n_t4, n_tier34=n_t34,
        n_t1_firm=n_t1_firm, n_t34_firm=n_t34_firm,
        spot_check_pct=round(spot_check_pct, 1),
        deep_review_pct=round(deep_review_pct, 1),
        t1_stats=t1_stats, t2_stats=t2_stats, t3_stats=t3_stats,
        t4_stats=t4_stats, t34_stats=t34_stats,
        t2_corrected=t2_corrected, t2_flagged=t2_flagged,
        t3_flagged=t3_flagged, t4_flagged=t4_flagged,
        n_cnn_errors_total=n_cnn_errors_total, n_pr_errors_total=n_pr_errors_total,
        n_ml_deferred=n_ml_deferred, n_ml_wrong_firm=n_ml_wrong_firm,
        baseline_density=round(baseline_density, 4),
        t1_error_density=round(t1_error_density, 4),
        t34_error_density=round(t34_error_density, 4),
        t2_error_density=round(t2_error_density, 4),
        t3_error_density=round(t3_error_density, 4),
        t4_error_density=round(t4_error_density, 4),
        density_ratio=round(density_ratio, 2) if density_ratio is not None else None,
        density_ratio_infinite=density_ratio_inf,
        triage_quality=triage_quality,
        t34_enrichment=round(t34_enrichment, 2) if t34_enrichment is not None else None,
        t34_enrichment_infinite=t34_enrichment_inf,
        t1_errors=t1_stats["n_remaining"],
        t1_error_rate_pct=round(t1_error_density * 100, 2),
        all_errors=all_errors_firm, ml_errors=n_ml_wrong_firm,
        t1_errors_count=t1_errors_firm,
        t2_errors=t2_errors_firm, t3_errors=t3_errors_firm,
        t4_errors=t4_errors_firm, t34_errors=t34_errors_firm,
        error_concentration_t34=round(firm_conc_t34, 1),
        error_in_t1_pct=round(firm_t1_pct, 1),
        all_problems=n_pr_errors_total,
        t1_problems=t1_stats["n_pr_errors"], t34_problems=t34_stats["n_pr_errors"],
        n_pr_deferred_all=n_pr_deferred_all,
        n_pr_deferred_t1=n_pr_deferred_t1, n_pr_deferred_t34=n_pr_deferred_t34,
        problem_concentration_t34=round(problem_concentration_t34, 1),
        problem_in_t1_pct=round(error_in_t1_pct, 1),
        nnr_ml=round(nnr_ml, 2) if nnr_ml != float("inf") else None,
        nnr_pr=round(nnr_pr, 2) if nnr_pr != float("inf") else None,
        nnr_improvement=nnr_improvement,
        nnr_ml_problems=nnr_ml_problems, nnr_pr_problems=nnr_pr_problems,
        n_r3_flagged=n_r3_flagged, n_r3_errors=n_r3_errors,
        r3_true_error_rate=round(r3_true_error_rate, 1),
        r3_false_alarm_rate=round(r3_false_alarm_rate, 1),
        n_flagged_total=n_flagged_total, n_flagged_errors=n_flagged_errors,
        n_flag_only=n_flag_only, n_flag_only_errors=n_flag_only_errors,
        flag_precision_all=round(flag_precision_all, 1),
        n_t2_improved=n_t2_improved, n_t2_worsened=n_t2_worsened,
        t2_correction_precision=t2_correction_precision,
    )


# ─────────────────────────────────────────────────────────────────
#  Main evaluate
# ─────────────────────────────────────────────────────────────────

def evaluate(
    raw: Dict, final: Dict, image_info: Dict,
    slides: Dict, slide_vote: Dict, slide_final: Dict,
    output_dir: str,
    image_info_all: Dict = None,
) -> Dict:
    os.makedirs(output_dir, exist_ok=True)

    for iname, fp in final.items():
        fp["gt_label"] = image_info[iname].get("gt_label")

    has_gt = any(fp.get("gt_label") is not None for fp in final.values())

    # ── Patch DataFrame ──────────────────────────────────────────
    rows = []
    for iname, fp in final.items():
        info    = image_info[iname]
        sv      = slide_vote.get(info["slide_id"], {})
        ml_pred = raw[iname]["predicted"]
        pr_pred = fp["final_predicted"]
        gt      = fp.get("gt_label")

        ml_ok = (ml_pred == gt) if gt else None
        pr_ok = (pr_pred == gt) if gt else None

        has_grid           = (info["row"] is not None)
        n_patches_in_slide = len(slides.get(info["slide_id"], []))

        if not has_grid:
            img_type = "individual"
        elif n_patches_in_slide == 1:
            img_type = "individual"
        elif n_patches_in_slide < config.MIN_PATCHES_FOR_VOTE:
            img_type = "sparse_grid"
        else:
            img_type = "grid"

        ml_zone = "firm"     if ml_pred in BINARY_LABELS else "deferred"
        pr_zone = "firm"     if pr_pred in BINARY_LABELS else "deferred"

        rows.append({
            "image_name":      iname,
            "slide_id":        info["slide_id"],
            "subfolder":       info["subfolder"],
            "filename":        info["filename"],
            "row":             info["row"],
            "col":             info["col"],
            "image_type":      img_type,
            "gt_label":        gt,
            "ml_predicted":    ml_pred,
            "ml_zone":         ml_zone,
            "ml_prob_G":       round(raw[iname]["prob_G"],     4),
            "ml_prob_Gplus":   round(raw[iname]["prob_Gplus"], 4),
            "ml_prob_Mixed":   round(max(0.0, 1.0 - raw[iname]["prob_G"] - raw[iname]["prob_Gplus"]), 4),
            "ml_conf":         round(raw[iname]["conf"],       4),
            "ml_conf_tier":    raw[iname]["conf_tier"],
            "ml_correct":      ml_ok,
            "final_predicted": pr_pred,
            "pr_zone":         pr_zone,
            "rule_applied":    fp["rule_applied"],
            "changed":         fp["changed"],
            "needs_review":    fp["needs_review"],
            "correction_confidence": fp.get("correction_confidence"),
            "pr_correct":      pr_ok,
            "outcome": (
                "improved" if (
                    (fp["changed"] or str(fp.get("rule_applied") or "").startswith(("rule1", "rule2")))
                    and (ml_ok is False) and (pr_ok is True)
                ) else
                "error" if (
                    (fp["changed"] or str(fp.get("rule_applied") or "").startswith(("rule1", "rule2")))
                    and (pr_ok is False)
                ) else
                "no_change"
            ) if gt else "unknown",
            "slide_majority":   sv.get("majority"),
            "slide_ratio":      round(sv.get("ratio", 0), 4),
            "slide_dominant":   sv.get("dominant"),
            "slide_standalone": sv.get("is_standalone", False),
            "slide_sparse":     sv.get("is_sparse", False),
        })

    patch_df = pd.DataFrame(rows)

    # ── Triage tier assignment ────────────────────────────────────
    _MEDIUM_THR = float(getattr(config, "MEDIUM_CONF_THR", 0.75))
    _HIGH_THR   = float(getattr(config, "HIGH_CONF_THR",   0.85))

    def _assign_tier(row):
        _changed  = bool(row["changed"])
        _rev      = bool(row["needs_review"])
        _conf     = float(row["ml_conf"])
        _is_mixed = str(row["ml_predicted"]) in UNCERTAIN_LABELS
        _rule     = str(row["rule_applied"] or "")

        if _changed or "rule1" in _rule or "rule2" in _rule:
            return 2
        if _rev and "rule3" in _rule:
            return 3
        if _rev and "rule4" in _rule:
            return 4
        if _rev and not _is_mixed and _conf >= _MEDIUM_THR:
            return 3
        if _rev or _is_mixed or _conf < _HIGH_THR:
            return 4
        return 1

    patch_df["triage_tier"]  = patch_df.apply(_assign_tier, axis=1)
    patch_df["triage_label"] = patch_df["triage_tier"].map(
        {1: "Confirmed", 2: "Suggested", 3: "Attention", 4: "Refer"}
    )

    def _suggested_label(row):
        pred = row.get("final_predicted")
        if pred not in BINARY_LABELS:
            return None
        tier = int(row.get("triage_tier", 0))
        rule = str(row.get("rule_applied") or "")
        if tier in (1, 2, 3) or "rule1" in rule or "rule2" in rule:
            return pred
        return None

    patch_df["suggested_label"] = patch_df.apply(_suggested_label, axis=1)
    _tc     = patch_df["triage_tier"].value_counts().sort_index()
    n_tier1 = int(_tc.get(1, 0))
    n_tier2 = int(_tc.get(2, 0))
    n_tier3 = int(_tc.get(3, 0))
    n_tier4 = int(_tc.get(4, 0))

    # ── Patch metrics ─────────────────────────────────────────────
    pm_ml, pm_pr  = {}, {}
    mixed_stats   = {}
    deferral_stats = {}
    fair_comparisons = {}

    if has_gt:
        binary_rows    = [r for r in rows if r["gt_label"] in BINARY_LABELS]
        deferral_stats = _compute_deferral_stats(binary_rows)

        ml_firm_rows = [r for r in binary_rows if r["ml_predicted"] in BINARY_LABELS]
        if ml_firm_rows:
            gt_i  = [_label_to_int(r["gt_label"])     for r in ml_firm_rows]
            ml_i  = [_label_to_int(r["ml_predicted"])  for r in ml_firm_rows]
            pm_ml = _cls_metrics(gt_i, ml_i)

        pr_firm_rows = [r for r in binary_rows if r["final_predicted"] in BINARY_LABELS]
        if pr_firm_rows:
            gt_i  = [_label_to_int(r["gt_label"])        for r in pr_firm_rows]
            pr_i  = [_label_to_int(r["final_predicted"])  for r in pr_firm_rows]
            pm_pr = _cls_metrics(gt_i, pr_i)

        fair_comparisons = _compute_fair_comparisons(rows)

        UNCERTAIN_GT = UNCERTAIN_LABELS
        info_source  = image_info_all if image_info_all else image_info
        all_mixed_gt = [n for n, info in info_source.items()
                        if info.get("gt_label") in UNCERTAIN_GT]
        if all_mixed_gt:
            n_mx            = len(all_mixed_gt)
            n_ml_flag       = sum(1 for n in all_mixed_gt if n in raw
                                  and raw[n]["conf_tier"] == "LOW")
            n_pr_flag       = sum(1 for n in all_mixed_gt if n in final
                                  and final[n].get("needs_review", False))
            n_ml_mixed_pred = sum(1 for n in all_mixed_gt if n in raw
                                  and raw[n]["predicted"] in UNCERTAIN_GT)
            n_changed_mx    = sum(1 for n in all_mixed_gt if n in final
                                  and final[n].get("changed", False))
            mixed_stats = dict(
                n_mixed=n_mx,
                n_ml_predicted_mixed=n_ml_mixed_pred,
                n_flagged_ml=n_ml_flag,
                n_flagged_pr=n_pr_flag,
                n_changed=n_changed_mx,
                flag_rate_ml=n_ml_flag / n_mx,
                flag_rate_pr=n_pr_flag / n_mx,
                ml_mixed_pred_rate=n_ml_mixed_pred / n_mx,
            )

    # ── Slide DataFrame ──────────────────────────────────────────
    slide_rows = []
    for sid, sf in slide_final.items():
        sv    = slide_vote[sid]
        ml_dg = sv["majority"] if sv["dominant"] else "UNCERTAIN"
        pr_dg = sf["diagnosis"]
        gt_dg = sf["slide_gt"]
        ml_ok = (ml_dg == gt_dg) if gt_dg and ml_dg not in ("UNCERTAIN",) else None
        pr_ok = (pr_dg == gt_dg) if gt_dg and pr_dg not in ("REFER", "TIE") else None
        pr_referred = (pr_dg == "REFER")
        slide_rows.append({
            "slide_id":       sid,
            "gt_label":       gt_dg,
            "ml_diagnosis":   ml_dg,
            "pr_diagnosis":   pr_dg,
            "pr_referred":    pr_referred,
            "ml_correct":     ml_ok,
            "pr_correct":     pr_ok,
            "outcome": (
                "referred"  if pr_referred else
                "improved"  if ml_ok is not None and not ml_ok and pr_ok else
                "worsened"  if ml_ok and pr_ok is not None and not pr_ok else
                "correct"   if ml_ok and pr_ok else
                "wrong"     if ml_ok is not None and not ml_ok and
                               pr_ok is not None and not pr_ok else "no_gt"
            ),
            "n_patches":      sf["total"],
            "n_G_final":      sf["n_G"],
            "n_Gplus_final":  sf["n_Gplus"],
            "n_Mixed_final":  sf.get("n_Mixed", 0),
            "n_changed":      sf["n_changed"],
            "n_needs_review": sf["n_needs_review"],
            "majority_ratio": round(sf["ratio"], 4),
            "dominant":       sf["dominant"],
        })
    slide_df = pd.DataFrame(slide_rows)

    # ── Slide metrics ─────────────────────────────────────────────
    sm_ml, sm_pr = {}, {}
    if has_gt and len(slide_df) > 0:
        s_gt    = slide_df["gt_label"].notna()
        valid_s = [
            (g, m, p)
            for g, m, p in zip(
                slide_df.loc[s_gt, "gt_label"],
                slide_df.loc[s_gt, "ml_diagnosis"],
                slide_df.loc[s_gt, "pr_diagnosis"],
            )
            if _label_to_int(g) is not None
            and _label_to_int(m) is not None
            and _label_to_int(p) is not None
            and p not in ("REFER", "TIE", "UNCERTAIN")
            and m not in ("UNCERTAIN",)
        ]
        if valid_s:
            g_s, m_s, p_s = zip(*valid_s)
            sm_ml = _cls_metrics([_label_to_int(v) for v in g_s],
                                  [_label_to_int(v) for v in m_s])
            sm_pr = _cls_metrics([_label_to_int(v) for v in g_s],
                                  [_label_to_int(v) for v in p_s])

    # ── Sub-DataFrames ───────────────────────────────────────────
    changed_df = patch_df[patch_df["changed"]].copy()
    review_df  = patch_df[patch_df["needs_review"]].copy()
    errors_df  = patch_df[
        (patch_df["pr_correct"] == False) &
        (patch_df["pr_zone"] == "firm") &
        patch_df["gt_label"].notna()
    ].copy() if has_gt else pd.DataFrame()

    # ── Coverage stats ───────────────────────────────────────────
    n_total      = len(patch_df)
    n_individual = int((patch_df["image_type"] == "individual").sum())
    n_sparse     = int((patch_df["image_type"] == "sparse_grid").sum())
    n_grid       = int((patch_df["image_type"] == "grid").sum())

    eligible_df  = patch_df[patch_df["image_type"] == "grid"]
    n_eligible   = len(eligible_df)
    n_pr_changed = int(eligible_df["changed"].sum())
    n_pr_flagged = int(eligible_df["needs_review"].sum())

    if has_gt:
        unflagged_errors_df = patch_df[
            (patch_df["pr_correct"] == False) &
            (patch_df["pr_zone"] == "firm") &
            (patch_df["needs_review"] == False) &
            patch_df["gt_label"].notna()
        ].copy()
        n_unflagged = len(unflagged_errors_df)
    else:
        unflagged_errors_df = pd.DataFrame()
        n_unflagged = None

    r3_df = patch_df[
        patch_df["rule_applied"].str.contains("rule3", na=False) & patch_df["needs_review"]
    ].copy()
    n_r3_flagged = len(r3_df)
    n_r3_errors  = int((r3_df["pr_correct"] == False).sum()) if (has_gt and len(r3_df) > 0) else 0

    r4a_df = patch_df[
        patch_df["rule_applied"].str.contains("rule4", na=False) & patch_df["needs_review"]
    ].copy()
    n_r4a_total       = len(r4a_df)
    n_r4a_errors      = int((r4a_df["ml_correct"] == False).sum()) if has_gt else None
    n_r4a_ok          = int((r4a_df["ml_correct"] == True).sum())  if has_gt else None
    n_r4a_false_alarm = n_r4a_ok if has_gt else None
    n_r4_total        = n_r4a_total
    n_r4_errors       = n_r4a_errors
    n_r4_false_alarm  = n_r4a_false_alarm

    saved_from_t3_df = patch_df[
        patch_df["correction_confidence"].notna()
    ].copy() if "correction_confidence" in patch_df.columns else pd.DataFrame()
    n_saved_strong   = int((saved_from_t3_df["correction_confidence"] == "strong").sum())   if len(saved_from_t3_df) else 0
    n_saved_moderate = int((saved_from_t3_df["correction_confidence"] == "moderate").sum()) if len(saved_from_t3_df) else 0
    n_saved_total    = n_saved_strong + n_saved_moderate
    if has_gt and n_saved_total > 0:
        n_saved_correct = int((saved_from_t3_df["pr_correct"] == True).sum())
    else:
        n_saved_correct = None

    workload_metrics = _compute_workload_metrics(
        patch_df, has_gt, n_r3_flagged, n_r3_errors
    ) if has_gt else {}

    def _type_metrics(df_sub):
        if not has_gt or len(df_sub) == 0:
            return {}
        firm_sub = df_sub[df_sub["pr_zone"] == "firm"]
        ml_ok_n  = int((firm_sub["ml_correct"] == True).sum())
        pr_ok_n  = int((firm_sub["pr_correct"] == True).sum())
        n_firm   = len(firm_sub)
        n_def    = int((df_sub["pr_zone"] == "deferred").sum())
        return dict(
            n=len(df_sub), n_firm=n_firm, n_deferred=n_def,
            ml_acc  = f"{ml_ok_n/n_firm*100:.1f}%" if n_firm else "N/A",
            pr_acc  = f"{pr_ok_n/n_firm*100:.1f}%" if n_firm else "N/A",
            delta   = f"{(pr_ok_n - ml_ok_n)/n_firm*100:+.1f}%" if n_firm else "N/A",
            changed = int(df_sub["changed"].sum()),
            improved= int((df_sub["outcome"]=="improved").sum()),
            worsened= int((df_sub["outcome"]=="worsened").sum()),
        )

    type_stats = {
        t: _type_metrics(patch_df[patch_df["image_type"] == t])
        for t in ["individual", "sparse_grid", "grid"]
    }

    # ── Save CSVs ────────────────────────────────────────────────
    patch_df.to_csv(f"{output_dir}/patch_results.csv",   index=False)
    slide_df.to_csv(f"{output_dir}/slide_results.csv",   index=False)
    if len(changed_df) > 0:
        changed_df.to_csv(f"{output_dir}/changed_patches.csv", index=False)

    # ── Console report ───────────────────────────────────────────
    _print_report(
        pm_ml, pm_pr, sm_ml, sm_pr,
        patch_df, changed_df, slide_df,
        n_individual, n_sparse, n_grid, n_eligible,
        n_r4_total, n_r4_errors, n_r4_false_alarm, n_unflagged,
        has_gt, output_dir,
        mixed_stats=mixed_stats,
        deferral_stats=deferral_stats,
        fair_comparisons=fair_comparisons,
        workload_metrics=workload_metrics,
    )

    # ── Excel workbook ───────────────────────────────────────────
    try:
        _save_excel(
            output_dir=output_dir,
            patch_df=patch_df,
            slide_df=slide_df,
            changed_df=changed_df,
            review_df=review_df,
            errors_df=errors_df,
            unflagged_errors_df=unflagged_errors_df,
            pm_ml=pm_ml, pm_pr=pm_pr,
            sm_ml=sm_ml, sm_pr=sm_pr,
            has_gt=has_gt,
            n_total=n_total, n_individual=n_individual,
            n_sparse=n_sparse, n_grid=n_grid, n_eligible=n_eligible,
            n_pr_changed=n_pr_changed, n_pr_flagged=n_pr_flagged,
            n_r4_total=n_r4_total, n_r4_errors=n_r4_errors,
            n_r4_false_alarm=n_r4_false_alarm, n_unflagged=n_unflagged,
            type_stats=type_stats,
            mixed_stats=mixed_stats,
            deferral_stats=deferral_stats,
            fair_comparisons=fair_comparisons,
            n_tier1=n_tier1, n_tier2=n_tier2, n_tier3=n_tier3, n_tier4=n_tier4,
            workload_metrics=workload_metrics,
            n_r3_flagged=n_r3_flagged, n_r3_errors=n_r3_errors,
            n_r4a_total=n_r4a_total, n_r4a_errors=n_r4a_errors,
            n_r4a_false_alarm=n_r4a_false_alarm,
            n_saved_total=n_saved_total, n_saved_strong=n_saved_strong,
            n_saved_moderate=n_saved_moderate, n_saved_correct=n_saved_correct,
        )
    except Exception as e:
        import traceback
        print(f"  [Excel] Could not save workbook: {e}")
        traceback.print_exc()

    return dict(
        patch_metrics_ml=pm_ml, patch_metrics_pr=pm_pr,
        slide_metrics_ml=sm_ml, slide_metrics_pr=sm_pr,
        deferral_stats=deferral_stats,
        fair_comparisons=fair_comparisons,
        mixed_stats=mixed_stats,
        has_gt=has_gt,
        n_patches=n_total, n_changed=int(patch_df["changed"].sum()),
        n_slides=len(slide_df),
        workload_metrics=workload_metrics,
    )


# ─────────────────────────────────────────────────────────────────
#  Excel
# ─────────────────────────────────────────────────────────────────

def _save_excel(
    output_dir, patch_df, slide_df, changed_df, review_df,
    errors_df, unflagged_errors_df,
    pm_ml, pm_pr, sm_ml, sm_pr, has_gt,
    n_total, n_individual, n_sparse, n_grid, n_eligible,
    n_pr_changed, n_pr_flagged,
    n_r4_total, n_r4_errors, n_r4_false_alarm, n_unflagged,
    type_stats, mixed_stats=None, deferral_stats=None,
    fair_comparisons=None, n_tier1=0, n_tier2=0, n_tier3=0, n_tier4=0,
    workload_metrics=None,
    n_r3_flagged=0, n_r3_errors=0,
    n_r4a_total=0, n_r4a_errors=None, n_r4a_false_alarm=None,
    n_saved_total=0, n_saved_strong=0, n_saved_moderate=0, n_saved_correct=None,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    deferral_stats   = deferral_stats   or {}
    mixed_stats      = mixed_stats      or {}
    fair_comparisons = fair_comparisons or {}

    wb = openpyxl.Workbook()

    # ── Style constants ───────────────────────────────────────────
    C_NAVY  = "1E3A5F"; C_BLUE  = "2E6DA4"
    C_GB    = "C6EFCE"; C_GT    = "276221"
    C_RB    = "FFCCCC"; C_RT    = "9C0006"
    C_AB    = "FFF2CC"; C_AT    = "7D6608"
    C_GREY  = "F2F2F2"; C_WHITE = "FFFFFF"
    C_PURP  = "E8E0F4"; C_PURPT = "4B0082"
    C_TEAL  = "D0F0F0"; C_TEALT = "005555"
    C_BLACK = "000000"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill(h):
        return PatternFill("solid", fgColor=str(h).replace("#", ""))

    def _hdr(ws, r, c, v, bg=C_NAVY, ft=C_WHITE, sz=10, bold=True):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, color=str(ft).replace("#", ""), size=sz)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        return cell

    def _cell(ws, r, c, v, bg=None, ft=C_BLACK, bold=False, fmt=None, align="center"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, color=str(ft).replace("#", ""), size=9)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = border
        if bg:  cell.fill          = _fill(bg)
        if fmt: cell.number_format = fmt
        return cell

    def _auto_w(ws, mn=8, mx=32):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, mn), mx)

    def _write_df(ws, df, sr=1, outcome_col=None, changed_col=None, correct_col=None):
        if df is None or len(df) == 0:
            ws.cell(row=sr, column=1, value="No data.")
            return
        for ci, col in enumerate(df.columns, 1):
            _hdr(ws, sr, ci, col)
        for ri, (_, row) in enumerate(df.iterrows(), sr+1):
            bg = C_GREY if ri % 2 == 0 else None
            oc = str(row.get(outcome_col, "")) if outcome_col else ""
            if   oc == "improved":             bg = C_GB
            elif oc in ("worsened", "error"):  bg = C_RB
            elif oc == "correct":              bg = C_GB
            elif oc == "wrong":                bg = C_RB
            elif oc == "referred":             bg = C_TEAL

            for ci, col in enumerate(df.columns, 1):
                val = row[col]
                cbg = bg; cft = C_BLACK; cbl = False

                if changed_col and col == changed_col and val is True:
                    cbg, cft, cbl = C_AB, C_AT, True
                if correct_col and col == correct_col:
                    if val is True:    cbg, cft = C_GB, C_GT
                    elif val is False: cbg, cft = C_RB, C_RT
                if col == "outcome":
                    if val == "improved":             cbg, cft, cbl = C_GB, C_GT, True
                    elif val in ("worsened", "error"): cbg, cft, cbl = C_RB, C_RT, True
                    elif val == "referred":            cbg, cft, cbl = C_TEAL, C_TEALT, True
                if col in ("ml_zone", "pr_zone"):
                    if val == "deferred": cbg, cft = C_TEAL, C_TEALT
                if col == "image_type":
                    if val == "individual":   cbg = C_AB
                    elif val == "sparse_grid": cbg = C_PURP

                fmt = "0.0000" if col in ("ml_conf", "ml_prob_G", "ml_prob_Gplus",
                                           "majority_ratio", "slide_ratio") else None
                _cell(ws, ri, ci, val, bg=cbg, ft=cft, bold=cbl, fmt=fmt)

    # ─────────────────────────────────────────────────────────────
    #  Sheet 1: Summary
    # ─────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    t = ws1.cell(row=1, column=1, value="Spatial Rule Engine — Evaluation Summary")
    t.font = Font(bold=True, size=14, color=C_WHITE)
    t.fill = _fill(C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells("A1:H1")
    ws1.row_dimensions[1].height = 28

    r = 3

    # ── TABLE B: Deferral / Coverage ──────────────────────────────
    ws1.cell(row=r, column=1,
             value="TABLE B — DEFERRAL & COVERAGE  (safety metric)").font = Font(
        bold=True, size=11, color=C_BLUE)
    r += 1
    note3 = ws1.cell(row=r, column=1,
        value="A Mixed prediction on a G/Gplus GT patch = system defers to human review. "
              "Higher deferral with good accuracy on firm calls = clinically desirable. "
              "'Coverage' = fraction of patches where a firm call was made.")
    note3.font      = Font(italic=True, size=9, color="595959")
    note3.alignment = Alignment(wrap_text=True, horizontal="left")
    ws1.merge_cells(f"A{r}:H{r}")
    r += 1

    # ── FIX: 4-column header (no Interpretation column) ──────────
    for ci, h in enumerate(["Metric", "CNN only", "Rule engine", "Change"], 1):
        _hdr(ws1, r, ci, h, bg=C_BLUE)
    r += 1

    ds      = deferral_stats
    n_tot_b = ds.get("n_total", 0)

    def _pct(num, denom):
        return f"{num/denom*100:.1f}%" if denom else "N/A"

    # ── FIX: 4-element tuples (no interp string) ─────────────────
    deferral_rows_b = [
        ("Total binary GT patches",
         n_tot_b, n_tot_b, "—"),
        ("Firm predictions (N)",
         ds.get("n_ml_firm", 0), ds.get("n_pr_firm", 0),
         f"{ds.get('n_pr_firm',0) - ds.get('n_ml_firm',0):+d}"),
        ("Coverage %",
         _pct(ds.get("n_ml_firm",0), n_tot_b),
         _pct(ds.get("n_pr_firm",0), n_tot_b),
         "—"),
        ("Deferred (Mixed prediction)",
         ds.get("n_ml_deferred", 0), ds.get("n_pr_deferred", 0),
         f"{ds.get('n_pr_deferred',0) - ds.get('n_ml_deferred',0):+d}"),
        ("Deferral rate %",
         _pct(ds.get("n_ml_deferred",0), n_tot_b),
         _pct(ds.get("n_pr_deferred",0), n_tot_b),
         "—"),
    ]
    # ── FIX: unpack 4 values, no interp ──────────────────────────
    for lbl, ml_v, pr_v, chg in deferral_rows_b:
        _cell(ws1, r, 1, lbl, bold=True, align="left")
        _cell(ws1, r, 2, ml_v, bg=C_AB)
        _cell(ws1, r, 3, pr_v, bg=C_TEAL)
        _cell(ws1, r, 4, chg)
        r += 1
    r += 1

    # ── TABLE C: Fair comparison ──────────────────────────────────
    if fair_comparisons:
        ws1.cell(row=r, column=1,
                 value="TABLE C — FAIR COMPARISON  (denominator analysis)").font = Font(
            bold=True, size=11, color=C_BLUE)
        r += 1
        note_fc = ws1.cell(row=r, column=1,
            value=(
                "Table A compares ML (N=ML-firm) vs PR (N=PR-firm). "
                "When PR resolves deferred patches, denominators differ — PR is tested on a larger, harder set. "
                "Scenario A is the fairest accuracy comparison. "
                "Scenario B shows the value of coverage. "
                "Scenario C shows quality of extra commits."
            ))
        note_fc.font      = Font(italic=True, size=9, color="595959")
        note_fc.alignment = Alignment(wrap_text=True, horizontal="left")
        ws1.merge_cells(f"A{r}:H{r}")
        ws1.row_dimensions[r].height = 36
        r += 1

        sa = fair_comparisons.get("scenario_A", {})
        if sa:
            ws1.cell(row=r, column=1,
                     value=f"Scenario A — Same N={sa.get('n',0)} (ML-firm patches only, apples-to-apples)").font = Font(
                bold=True, size=10, color=C_BLUE)
            r += 1
            for ci, h in enumerate(["Metric", "ML only (N=ML-firm)", "PR (same N)", "Delta", "Interpretation"], 1):
                _hdr(ws1, r, ci, h, bg=C_BLUE)
            r += 1
            ml_a = sa.get("ml", {}); pr_a = sa.get("pr", {})
            for key, label in [("mcc","MCC"), ("bal_acc","Balanced Accuracy"),
                                ("recall","Recall"), ("precision","Precision"), ("f1","F1")]:
                ml_v  = float(ml_a.get(key, 0))
                pr_v  = float(pr_a.get(key, 0))
                delta = pr_v - ml_v
                bg    = C_GB if delta > 0.001 else (C_RB if delta < -0.001 else None)
                ft    = C_GT if delta > 0.001 else (C_RT if delta < -0.001 else "000000")
                interp = (
                    "PR improved on ML committed patches" if delta > 0.001 else
                    "PR slightly harmed ML committed patches" if delta < -0.001 else
                    "No change on shared patches"
                )
                _cell(ws1, r, 1, label, bold=True, align="left")
                _cell(ws1, r, 2, round(ml_v, 4), fmt="0.0000", bg=C_AB)
                _cell(ws1, r, 3, round(pr_v, 4), fmt="0.0000", bg=bg, ft=ft, bold=(delta > 0.001))
                _cell(ws1, r, 4, round(delta, 4), fmt="+0.0000;-0.0000", bg=bg, ft=ft)
                _cell(ws1, r, 5, interp, align="left")
                r += 1
            _cell(ws1, r, 1,
                  f"N(PR deferred on ML-firm set) = {sa.get('n_pr_deferred_on_ml_set',0)}",
                  align="left", bg=C_TEAL, ft=C_TEALT)
            r += 2

        sb = fair_comparisons.get("scenario_B", {})
        if sb:
            ws1.cell(row=r, column=1,
                     value=f"Scenario B — All {sb.get('n',0)} patches, Mixed = forced wrong").font = Font(
                bold=True, size=10, color=C_BLUE)
            r += 1
            for ci, h in enumerate(["Metric", "ML (forced, N=all)", "PR (forced, N=all)", "Delta", "Interpretation"], 1):
                _hdr(ws1, r, ci, h, bg=C_BLUE)
            r += 1
            ml_b = sb.get("ml", {}); pr_b = sb.get("pr", {})
            for key, label in [("mcc","MCC"), ("bal_acc","Balanced Accuracy"),
                                ("recall","Recall"), ("f1","F1")]:
                ml_v  = float(ml_b.get(key, 0))
                pr_v  = float(pr_b.get(key, 0))
                delta = pr_v - ml_v
                bg    = C_GB if delta > 0.001 else (C_RB if delta < -0.001 else None)
                ft    = C_GT if delta > 0.001 else (C_RT if delta < -0.001 else "000000")
                interp = ("Significant improvement" if delta > 0.02 else
                          "Modest improvement"      if delta > 0.001 else
                          "No change / slight decline")
                _cell(ws1, r, 1, label, bold=True, align="left")
                _cell(ws1, r, 2, round(ml_v, 4), fmt="0.0000", bg=C_AB)
                _cell(ws1, r, 3, round(pr_v, 4), fmt="0.0000", bg=bg, ft=ft, bold=(delta > 0.001))
                _cell(ws1, r, 4, round(delta, 4), fmt="+0.0000;-0.0000", bg=bg, ft=ft)
                _cell(ws1, r, 5, interp, align="left")
                r += 1
            r += 1

        sc = fair_comparisons.get("scenario_C", {})
        if sc and sc.get("n", 0) > 0:
            ws1.cell(row=r, column=1,
                     value=f"Scenario C — Extra {sc.get('n',0)} patches PR committed (ML deferred, PR firm)").font = Font(
                bold=True, size=10, color=C_BLUE)
            r += 1
            for ci, h in enumerate(["Metric", "Value", "Interpretation"], 1):
                _hdr(ws1, r, ci, h, bg=C_BLUE)
            r += 1
            n_sc   = sc.get("n", 0)
            acc_sc = sc.get("accuracy", 0)
            bg_acc = C_GB if acc_sc >= 0.75 else (C_AB if acc_sc >= 0.60 else C_RB)
            sc_rows = [
                ("Extra patches committed",     n_sc,
                 "ML deferred, PR gave firm answer"),
                ("Correct (PR right)",          sc.get("n_correct", 0),
                 f"{sc.get('n_correct',0)/n_sc*100:.1f}% of extra" if n_sc else "—"),
                ("Wrong (PR introduced error)", sc.get("n_wrong", 0),
                 f"{sc.get('n_wrong',0)/n_sc*100:.1f}% of extra" if n_sc else "—"),
                ("Accuracy on extra patches",   f"{acc_sc*100:.1f}%",
                 "≥75% = good; <60% = net harmful coverage"),
                ("MCC on extra patches",
                 round(sc.get("metrics",{}).get("mcc",0), 4),
                 "Binary metric on the extra-commit set"),
            ]
            for lbl, val, interp in sc_rows:
                _cell(ws1, r, 1, lbl, bold=True, align="left")
                _cell(ws1, r, 2, val, bg=bg_acc if "Accuracy" in lbl else C_AB)
                _cell(ws1, r, 3, interp, align="left")
                r += 1
            r += 1

    # ── Confusion matrix ──────────────────────────────────────────
    ws1.cell(row=r, column=1,
             value="CONFUSION MATRIX  (firm predictions only)").font = Font(
        bold=True, size=11, color=C_BLUE)
    r += 1
    note4 = ws1.cell(row=r, column=1,
        value="Positive class = Gplus. Deferred patches excluded.")
    note4.font      = Font(italic=True, size=9, color="595959")
    note4.alignment = Alignment(wrap_text=True, horizontal="left")
    ws1.merge_cells(f"A{r}:H{r}")
    r += 1
    for ci, h in enumerate(["", "TP", "TN", "FP", "FN", "N (firm)"], 1):
        _hdr(ws1, r, ci, h, bg=C_BLUE)
    r += 1
    for label, pm, is_pr in [("CNN only", pm_ml, False), ("Rule engine", pm_pr, True)]:
        tp_v     = int(pm.get("tp", 0));  tn_v = int(pm.get("tn", 0))
        fp_v     = int(pm.get("fp", 0));  fn_v = int(pm.get("fn", 0))
        n_firm_v = tp_v + tn_v + fp_v + fn_v
        _cell(ws1, r, 1, label,    bold=is_pr, align="left")
        _cell(ws1, r, 2, tp_v,     bg=C_GB, bold=is_pr)
        _cell(ws1, r, 3, tn_v,     bg=C_GB, bold=is_pr)
        _cell(ws1, r, 4, fp_v,     bg=C_RB)
        _cell(ws1, r, 5, fn_v,     bg=C_RB)
        _cell(ws1, r, 6, n_firm_v, bold=is_pr)
        r += 1
    r += 1

    # ── Slide-level metrics ───────────────────────────────────────
    if sm_ml:
        ws1.cell(row=r, column=1,
                 value="SLIDE-LEVEL METRICS  (firm diagnoses only — REFER excluded)").font = Font(
            bold=True, size=11, color=C_BLUE)
        r += 1
        for ci, h in enumerate(["Metric", "CNN only", "Rule engine", "Delta", "% Change"], 1):
            _hdr(ws1, r, ci, h, bg=C_BLUE)
        r += 1
        for key, label in [("mcc","MCC"), ("bal_acc","Balanced Accuracy"),
                            ("f1","F1"), ("f2","F2"),
                            ("acc","Accuracy"), ("precision","Precision"), ("recall","Recall")]:
            ml_v  = float(sm_ml.get(key, 0)); pr_v = float(sm_pr.get(key, 0))
            delta = pr_v - ml_v; pct = (delta / ml_v * 100) if ml_v else 0.0
            bg = C_GB if delta > 0.001 else (C_RB if delta < -0.001 else None)
            ft = C_GT if delta > 0.001 else (C_RT if delta < -0.001 else "000000")
            _cell(ws1, r, 1, label, bold=True, align="left")
            _cell(ws1, r, 2, round(ml_v, 4), fmt="0.0000", bg=C_AB)
            _cell(ws1, r, 3, round(pr_v, 4), fmt="0.0000", bg=bg, ft=ft)
            _cell(ws1, r, 4, round(delta, 4), fmt="+0.0000;-0.0000", bg=bg, ft=ft)
            _cell(ws1, r, 5, f"{pct:+.2f}%", bg=bg, ft=ft)
            r += 1
        n_referred = int((slide_df["pr_diagnosis"] == "REFER").sum()) if len(slide_df) else 0
        _cell(ws1, r, 1,
              f"Slides referred (REFER): {n_referred} ({n_referred/len(slide_df)*100:.1f}%)"
              if len(slide_df) else "No slides",
              align="left", bg=C_TEAL, ft=C_TEALT)
        r += 2

    # ── Rule breakdown ────────────────────────────────────────────
    ws1.cell(row=r, column=1, value="RULE BREAKDOWN").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    _n_r3_flag  = int(patch_df[
        patch_df["rule_applied"].str.contains("rule3", na=False) & patch_df["needs_review"]
    ].shape[0])
    _n_r4a_flag = int(patch_df[
        patch_df["rule_applied"].str.contains("rule4", na=False) & patch_df["needs_review"]
    ].shape[0])

    n_ch = int(patch_df["changed"].sum())
    for ci, h in enumerate(["Rule", "Total", "Improved", "Worsened", "No GT change", "Tier", "Type"], 1):
        _hdr(ws1, r, ci, h, bg=C_BLUE)
    r += 1

    n_imp_all = int((patch_df[patch_df["changed"]]["outcome"] == "improved").sum()) if n_ch else 0
    n_wor_all = int((patch_df[patch_df["changed"]]["outcome"] == "error").sum())    if n_ch else 0
    n_sam_all = n_ch - n_imp_all - n_wor_all
    _cell(ws1, r, 1, "All spatial corrections (total)", bold=True, align="left")
    _cell(ws1, r, 2, n_ch,      bg=C_AB, bold=True)
    _cell(ws1, r, 3, n_imp_all, bg=C_GB, ft=C_GT, bold=True)
    _cell(ws1, r, 4, n_wor_all, bg=C_RB if n_wor_all else None,
          ft=C_RT if n_wor_all else C_BLACK, bold=bool(n_wor_all))
    _cell(ws1, r, 5, n_sam_all, bg=C_GREY if has_gt else C_AB)
    _cell(ws1, r, 6, "Tier 2")
    _cell(ws1, r, 7, "prediction changed", align="left")
    r += 1

    if n_ch > 0 and has_gt:
        for rule in patch_df[patch_df["changed"]]["rule_applied"].dropna().unique():
            sub_r = patch_df[patch_df["changed"] & (patch_df["rule_applied"] == rule)]
            cnt   = len(sub_r)
            imp   = int((sub_r["outcome"] == "improved").sum())
            wor   = int((sub_r["outcome"] == "error").sum())
            sam   = cnt - imp - wor
            _cell(ws1, r, 1, f"  {rule}", align="left")
            _cell(ws1, r, 2, cnt, bg=C_AB)
            _cell(ws1, r, 3, imp, bg=C_GB if imp else None, ft=C_GT if imp else C_BLACK)
            _cell(ws1, r, 4, wor, bg=C_RB if wor else None, ft=C_RT if wor else C_BLACK, bold=bool(wor))
            _cell(ws1, r, 5, sam, bg=C_GREY if sam else None)
            _cell(ws1, r, 6, "Tier 2")
            _cell(ws1, r, 7, "changed", align="left")
            r += 1

    _cell(ws1, r, 1, "rule3_cluster_flag (Tier 3)", align="left")
    _cell(ws1, r, 2, _n_r3_flag, bg=C_TEAL, ft=C_TEALT)
    _cell(ws1, r, 3, "—"); _cell(ws1, r, 4, "—"); _cell(ws1, r, 5, "—")
    _cell(ws1, r, 6, "Tier 3", bg=C_TEAL, ft=C_TEALT)
    _cell(ws1, r, 7, "Cluster flag → spatial attention zone", align="left")
    r += 1

    _cell(ws1, r, 1, "rule4_refer (Tier 4)", align="left")
    _cell(ws1, r, 2, _n_r4a_flag, bg=C_RB, ft=C_RT)
    _cell(ws1, r, 3, "—"); _cell(ws1, r, 4, "—"); _cell(ws1, r, 5, "—")
    _cell(ws1, r, 6, "Tier 4", bg=C_RB, ft=C_RT)
    _cell(ws1, r, 7, "Isolated LOW/Mixed → full human decision", align="left")
    r += 1

    if n_saved_total > 0:
        _cell(ws1, r, 1, "  LOW/Mixed saved by Rule 1/2 → Tier 2", align="left")
        _cell(ws1, r, 2, n_saved_total, bg=C_GB, ft=C_GT, bold=True)
        _cell(ws1, r, 3, n_saved_correct if n_saved_correct is not None else "—",
              bg=C_GB if n_saved_correct else None)
        _cell(ws1, r, 4, "—"); _cell(ws1, r, 5, "—")
        _cell(ws1, r, 6, "Tier 2", bg=C_GB, ft=C_GT)
        _cell(ws1, r, 7,
              f"strong={n_saved_strong} (Rule 1), moderate={n_saved_moderate} (Rule 2)",
              align="left")
        r += 1

    _auto_w(ws1)


    # ─────────────────────────────────────────────────────────────
    #  Sheet: Workload
    # ─────────────────────────────────────────────────────────────
    ws_wl = wb.create_sheet("Workload")
    ws_wl.sheet_view.showGridLines = False
    t_wl = ws_wl.cell(row=1, column=1,
        value="Triage Workload Analysis — Primary Clinical Metric")
    t_wl.font      = Font(bold=True, size=14, color=C_WHITE)
    t_wl.fill      = _fill(C_NAVY)
    t_wl.alignment = Alignment(horizontal="center", vertical="center")
    ws_wl.merge_cells("A1:I1")
    ws_wl.row_dimensions[1].height = 30

    wm = workload_metrics or {}
    r = 3

    ws_wl.cell(row=r, column=1,
               value="A — TIER DISTRIBUTION").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Tier", "Label", "N patches", "% of Total",
                             "N firm patches", "Role"], 1):
        _hdr(ws_wl, r, ci, h, bg=C_BLUE)
    r += 1
    tier_rows_wl = [
        (1,  "Confirmed", wm.get("n_tier1",  0), wm.get("spot_check_pct",  0), wm.get("n_t1_firm", 0),
         "HIGH-conf, firm, no rule fired. Spot-check depth."),
        (2,  "Corrected", wm.get("n_tier2",  0), "—", "—",
         "Rule corrected or Rule 1/2 spatially supported. Suggested label shown."),
        (3,  "Attention", wm.get("n_tier3",  0), "—", "—",
         "Cluster flag — uncertain zone, confirm collectively."),
        (4,  "Refer",     wm.get("n_tier4",  0), "—", "—",
         "Isolated LOW/Mixed. No suggested label. Full human decision."),
        (34, "Tier 3+4",  wm.get("n_tier34", 0), wm.get("deep_review_pct", 0), wm.get("n_t34_firm", 0),
         "Deep-review pool — Attention + Refer combined."),
    ]
    tier_bg_map = {1: C_GB, 2: C_AB, 3: C_AB, 4: C_RB, 34: C_PURP}
    tier_ft_map = {1: C_GT, 2: C_AT, 3: C_AT, 4: C_RT, 34: C_PURPT}
    for tier, label, n_p, pct, n_f, role in tier_rows_wl:
        bg = tier_bg_map[tier]; ft = tier_ft_map[tier]
        pct_str = f"{pct:.1f}%" if isinstance(pct, float) else str(pct)
        _cell(ws_wl, r, 1, tier,    bg=bg, ft=ft, bold=True)
        _cell(ws_wl, r, 2, label,   bg=bg, ft=ft, bold=True, align="left")
        _cell(ws_wl, r, 3, n_p,     bg=bg, ft=ft, bold=True)
        _cell(ws_wl, r, 4, pct_str, bg=bg, ft=ft)
        _cell(ws_wl, r, 5, n_f)
        _cell(ws_wl, r, 6, role, align="left")
        r += 1
    r += 1

    ws_wl.cell(row=r, column=1,
               value="B — ERROR DENSITY PER TIER").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Tier", "N patches", "N errors (CNN)",
                             "N errors (Rule-based)", "Handled by rule",
                             "Remaining silent error", "Density rate",
                             "Interpretation"], 1):
        _hdr(ws_wl, r, ci, h, bg=C_BLUE)
    r += 1

    t1s  = wm.get("t1_stats",  {})
    t2s  = wm.get("t2_stats",  {})
    t3s  = wm.get("t3_stats",  {})
    t4s  = wm.get("t4_stats",  {})
    t34s = wm.get("t34_stats", {})
    t2_corr     = wm.get("t2_corrected", 0)
    t2_flag_wl  = wm.get("t2_flagged", 0)
    t2_hand_str = (
        f"{t2s.get('n_handled', 0)} ({t2_corr} corrected"
        + (f", {t2_flag_wl} flagged" if t2_flag_wl else "") + ")"
        if t2s.get("n_cnn_errors", 0) > 0 else "0"
    )
    t3_flag_wl  = wm.get("t3_flagged", 0)
    n_cnn_total = wm.get("n_cnn_errors_total", 0)

    tier_table_rows = [
        ("Baseline (CNN, no triage)",
         wm.get("n_total", 0), n_cnn_total, n_cnn_total, "—", n_cnn_total,
         f"{n_cnn_total/wm.get('n_total',1)*100:.1f}%" if wm.get("n_total") else "—",
         "Uniform — all errors distributed across all patches", C_GREY, C_BLACK),
        ("Tier 1  Confirmed",
         t1s.get("n",0), t1s.get("n_cnn_errors",0), t1s.get("n_pr_errors",0),
         "0 (no rules fire on Tier 1)", t1s.get("n_remaining",0),
         f"{t1s.get('density_rate',0):.1f}%", "HIGH-conf anchors — should be near 0%", C_GB, C_GT),
        ("Tier 2  Suggested",
         t2s.get("n",0), t2s.get("n_cnn_errors",0), t2s.get("n_pr_errors",0),
         t2_hand_str, t2s.get("n_remaining",0),
         f"{t2s.get('density_rate',0):.1f}%",
         "Rule-supported; errors can decrease after correction", C_AB, C_AT),
        ("Tier 3  Attention",
         t3s.get("n",0), t3s.get("n_cnn_errors",0), t3s.get("n_pr_errors",0),
         f"{wm.get('t3_flagged',0)} flagged (cluster)", t3s.get("n_remaining",0),
         f"{t3s.get('density_rate',0):.1f}%",
         "Cluster flag — problems routed to review", C_TEAL, C_TEALT),
        ("Tier 4  Refer",
         t4s.get("n",0), t4s.get("n_cnn_errors",0), t4s.get("n_pr_errors",0),
         f"{wm.get('t4_flagged',0)} flagged (isolated)", t4s.get("n_remaining",0),
         f"{t4s.get('density_rate',0):.1f}%",
         "Isolated LOW/Mixed — problems routed to referral", C_RB, C_RT),
        ("Tier 3+4  Deep-review pool",
         t34s.get("n",0), t34s.get("n_cnn_errors",0), t34s.get("n_pr_errors",0),
         f"{t34s.get('n_handled',0)}", t34s.get("n_remaining",0),
         f"{t34s.get('density_rate',0):.1f}%",
         "Combined review pool — post-rule errors concentrated here", C_PURP, C_PURPT),
    ]
    for (lbl, n_p, n_cnn, n_pr, hand, remain, rate, interp, bg, ft) in tier_table_rows:
        is_t34 = "3+4" in lbl
        _cell(ws_wl, r, 1, lbl,    bg=bg, ft=ft, bold=("Baseline" not in lbl), align="left")
        _cell(ws_wl, r, 2, n_p,    bg=bg, ft=ft)
        _cell(ws_wl, r, 3, n_cnn,  bg=bg, ft=ft)
        _cell(ws_wl, r, 4, n_pr,   bg=bg, ft=ft, bold=is_t34)
        _cell(ws_wl, r, 5, hand,   align="left")
        _cell(ws_wl, r, 6, remain,
              bg=(C_RB if remain > 0 and "Baseline" not in lbl else bg),
              ft=(C_RT if remain > 0 and "Baseline" not in lbl else ft), bold=is_t34)
        _cell(ws_wl, r, 7, rate,   bold=True)
        _cell(ws_wl, r, 8, interp, align="left")
        r += 1
    r += 1

    dr      = wm.get("density_ratio")
    dr_inf  = wm.get("density_ratio_infinite", False)
    dr_str  = "∞ (Tier 1 clean)" if dr_inf else (f"{dr:.2f}" if dr is not None else "N/A")
    tq      = wm.get("triage_quality", "N/A")
    tq_bg   = C_GB if tq == "strong" else (C_AB if tq == "adequate" else C_RB)
    tq_ft   = C_GT if tq == "strong" else (C_AT if tq == "adequate" else C_RT)
    enrich  = wm.get("t34_enrichment")
    for ci, h in enumerate(["Metric", "Value", "Threshold", "Interpretation"], 1):
        _hdr(ws_wl, r, ci, h, bg=C_BLUE)
    r += 1
    summary_wl = [
        ("Density ratio (Tier3+4 / Tier1 error density)", dr_str,
         "≥3.0 strong, ≥1.5 adequate, <1.0 failed",
         "Primary safety check"),
        ("Triage quality", tq, "strong / adequate / weak / failed",
         "Overall rating based on density ratio"),
        ("Tier3+4 enrichment vs CNN uniform baseline",
         f"{enrich:.2f}x" if enrich else "∞",
         ">1.0 = review pool more error-dense than uniform",
         "Gain over no-triage strategy"),
        ("Problem concentration in Tier 3+4",
         f"{wm.get('error_concentration_t34',0):.1f}%",
         ">80% = most errors routed to review pool",
         "% of all errors in Tier 3+4"),
        ("Tier 1 error rate",
         f"{wm.get('t1_error_rate_pct',0):.2f}%",
         "<2% ideal",
         "% of Tier 1 patches still wrong"),
    ]
    for lbl, val, thr, interp in summary_wl:
        is_tq = "Triage quality" in lbl
        _cell(ws_wl, r, 1, lbl, bold=True, align="left")
        _cell(ws_wl, r, 2, val, bg=tq_bg if is_tq else C_AB,
              ft=tq_ft if is_tq else C_BLACK, bold=is_tq)
        _cell(ws_wl, r, 3, thr,    align="left")
        _cell(ws_wl, r, 4, interp, align="left")
        r += 1
    r += 1

    ws_wl.cell(row=r, column=1,
               value="C — NUMBER NEEDED TO REVIEW (NNR)").font = Font(
        bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["System", "Patches Reviewed", "Errors in Pool",
                             "NNR", "Improvement", "Interpretation"], 1):
        _hdr(ws_wl, r, ci, h, bg=C_BLUE)
    r += 1
    nnr_ml  = wm.get("nnr_ml");  nnr_pr  = wm.get("nnr_pr")
    nnr_imp = wm.get("nnr_improvement")
    _cell(ws_wl, r, 1, "CNN only (uniform review)", bold=True, align="left", bg=C_GREY)
    _cell(ws_wl, r, 2, wm.get("n_total",0),         bg=C_GREY)
    _cell(ws_wl, r, 3, wm.get("nnr_ml_problems",0), bg=C_GREY)
    _cell(ws_wl, r, 4, f"{nnr_ml:.2f}" if nnr_ml else "N/A", bg=C_GREY, bold=True)
    _cell(ws_wl, r, 5, "1.0× (baseline)", bg=C_GREY)
    _cell(ws_wl, r, 6, "No triage — expert reviews every patch equally", align="left")
    r += 1
    nnr_bg = C_GB if (nnr_ml and nnr_pr and nnr_pr < nnr_ml) else C_RB
    _cell(ws_wl, r, 1, "Rule engine — Tier 3+4 only", bold=True, align="left", bg=C_GB)
    _cell(ws_wl, r, 2, wm.get("n_tier34",0),         bg=C_GB)
    _cell(ws_wl, r, 3, wm.get("nnr_pr_problems",0),  bg=C_GB)
    _cell(ws_wl, r, 4, f"{nnr_pr:.2f}" if nnr_pr else "N/A", bg=C_GB, ft=C_GT, bold=True)
    _cell(ws_wl, r, 5,
          f"{nnr_imp:.1f}× faster" if nnr_imp else "N/A",
          bg=C_GB if (nnr_imp and nnr_imp > 1) else C_RB,
          ft=C_GT if (nnr_imp and nnr_imp > 1) else C_RT, bold=True)
    _cell(ws_wl, r, 6, "Expert reviews Tier 3+4 only — finds errors faster per patch", align="left")
    r += 2

    ws_wl.cell(row=r, column=1,
               value="D — TIER 2 CORRECTION QUALITY").font = Font(
        bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Metric", "Count / Value", "Interpretation"], 1):
        _hdr(ws_wl, r, ci, h, bg=C_BLUE)
    r += 1
    t2_cp = wm.get("t2_correction_precision", 0)
    t2_bg = C_GB if t2_cp >= 80 else (C_AB if t2_cp >= 60 else C_RB)
    t2_ft = C_GT if t2_cp >= 80 else (C_AT if t2_cp >= 60 else C_RT)
    d_rows = [
        ("Tier 2 patches corrected by spatial rule",
         wm.get("n_t2_improved",0) + wm.get("n_t2_worsened",0),
         "Spatial context rule changed the label"),
        ("  Improved (correction fixed a CNN error)",
         wm.get("n_t2_improved",0), "CNN was wrong, spatial rule corrected it"),
        ("  Worsened (correction introduced error)",
         wm.get("n_t2_worsened",0), "CNN was right, spatial rule broke it"),
        ("Correction precision", f"{t2_cp:.1f}%",
         "improved / (improved + error). ≥80% = reliable."),
        ("LOW/Mixed promoted from Tier 4 (total)",
         n_saved_total, "Patches corrected by spatial rule → Tier 2 instead of Tier 4"),
        ("  Promoted by dominant context rule (strong)",
         n_saved_strong, "High-evidence correction"),
        ("  Promoted by neighbour agreement (moderate)",
         n_saved_moderate, "Local 4-neighbour consensus"),
        ("  Promoted patches correct after correction",
         n_saved_correct if n_saved_correct is not None else "N/A",
         f"Accuracy on promoted patches: "
         f"{n_saved_correct/n_saved_total*100:.1f}%" if (n_saved_correct and n_saved_total) else "—"),
    ]
    for lbl, val, interp in d_rows:
        is_prec = "precision" in lbl.lower()
        is_wor  = "Worsened" in lbl
        bg = t2_bg if is_prec else (C_RB if is_wor and isinstance(val, int) and val > 0 else None)
        ft = t2_ft if is_prec else (C_RT if is_wor and isinstance(val, int) and val > 0 else C_BLACK)
        _cell(ws_wl, r, 1, lbl, bold=is_prec, align="left")
        _cell(ws_wl, r, 2, val, bg=bg, ft=ft, bold=is_prec)
        _cell(ws_wl, r, 3, interp, align="left")
        r += 1

    _auto_w(ws_wl)
    ws_wl.column_dimensions["A"].width = 42
    ws_wl.column_dimensions["H"].width = 48

    # ─────────────────────────────────────────────────────────────
    #  Sheet: Coverage
    # ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Coverage")
    ws2.sheet_view.showGridLines = False
    t2_cell = ws2.cell(row=1, column=1, value="Image Coverage & Rule Engine Eligibility")
    t2_cell.font      = Font(bold=True, size=13, color=C_WHITE)
    t2_cell.fill      = _fill(C_NAVY)
    t2_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.merge_cells("A1:F1")
    ws2.row_dimensions[1].height = 24

    r = 3
    ws2.cell(row=r, column=1, value="IMAGE TYPE BREAKDOWN").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Type", "Count", "% of Total", "Can rules fire?", "Explanation"], 1):
        _hdr(ws2, r, ci, h, bg=C_BLUE)
    r += 1
    type_info = [
        ("individual",  n_individual, "NO",
         "No row/col in filename OR only 1 patch in slide."),
        ("sparse_grid", n_sparse,     "NO",
         f"Has row/col but fewer than {config.MIN_PATCHES_FOR_VOTE} firm patches."),
        ("grid",        n_grid,       "YES",
         "Has row/col AND enough patches. Full rule-eligible."),
    ]
    type_bg_cov = {"individual": C_AB, "sparse_grid": C_PURP, "grid": C_GB}
    type_ft_cov = {"individual": C_AT, "sparse_grid": C_PURPT, "grid": C_GT}
    for tname, cnt, can_fire, expl in type_info:
        pct = f"{cnt/n_total*100:.1f}%" if n_total else "0%"
        bg  = type_bg_cov[tname]; ft = type_ft_cov[tname]
        _cell(ws2, r, 1, tname,    bg=bg, ft=ft, bold=True, align="left")
        _cell(ws2, r, 2, cnt,      bg=bg, ft=ft)
        _cell(ws2, r, 3, pct,      bg=bg, ft=ft)
        _cell(ws2, r, 4, can_fire, bg=(C_GB if can_fire == "YES" else C_RB),
              ft=(C_GT if can_fire == "YES" else C_RT), bold=True)
        _cell(ws2, r, 5, expl, align="left")
        r += 1
    _cell(ws2, r, 1, "TOTAL", bold=True, align="left")
    _cell(ws2, r, 2, n_total, bold=True)
    _cell(ws2, r, 3, "100%", bold=True)
    r += 2

    if deferral_stats and deferral_stats.get("n_total", 0) > 0:
        ws2.cell(row=r, column=1,
                 value="DEFERRAL BREAKDOWN").font = Font(bold=True, size=11, color=C_BLUE)
        r += 1
        for ci, h in enumerate(["Metric", "CNN only", "Rule engine", "Explanation"], 1):
            _hdr(ws2, r, ci, h, bg=C_BLUE)
        r += 1
        ds_cov = deferral_stats
        def_rows_cov = [
            ("Total G/Gplus GT patches",
             ds_cov.get("n_total",0), ds_cov.get("n_total",0), "Denominator for all rates"),
            ("Firm predictions",
             ds_cov.get("n_ml_firm",0), ds_cov.get("n_pr_firm",0),
             "G or Gplus output — used in Table A metrics"),
            ("Coverage %",
             f"{ds_cov.get('ml_coverage',0)*100:.1f}%",
             f"{ds_cov.get('pr_coverage',0)*100:.1f}%",
             "Firm / Total"),
            ("Deferred (Mixed output)",
             ds_cov.get("n_ml_deferred",0), ds_cov.get("n_pr_deferred",0),
             "Routed to human review — safe outcome"),
            ("Deferral rate %",
             f"{ds_cov.get('ml_deferral_rate',0)*100:.1f}%",
             f"{ds_cov.get('pr_deferral_rate',0)*100:.1f}%",
             "Deferred / Total"),
        ]
        for lbl, ml_v, pr_v, expl in def_rows_cov:
            _cell(ws2, r, 1, lbl, bold=True, align="left")
            _cell(ws2, r, 2, ml_v, bg=C_AB)
            _cell(ws2, r, 3, pr_v, bg=C_TEAL, ft=C_TEALT)
            _cell(ws2, r, 4, expl, align="left")
            r += 1
        r += 1

    ws2.cell(row=r, column=1,
             value="RULE ENGINE FIRING SUMMARY (grid patches only)").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Metric", "Count", "% of Grid"], 1):
        _hdr(ws2, r, ci, h, bg=C_BLUE)
    r += 1
    for label, cnt in [
        ("Grid patches (eligible)",          n_grid),
        ("Changed by Rule 1 or 2  (Tier 2)", n_pr_changed),
        ("Flagged by Rule 3  (Tier 3)",       n_r3_flagged),
        ("Flagged by Rule 4  (Tier 4)",       n_r4a_total),
    ]:
        pct = f"{cnt/n_grid*100:.1f}%" if n_grid else "0%"
        bg  = (C_GB if "Changed" in label else
               C_AB if "Rule 3" in label else
               C_RB if "Rule 4" in label else None)
        _cell(ws2, r, 1, label, bold=True, align="left")
        _cell(ws2, r, 2, cnt, bg=bg)
        _cell(ws2, r, 3, pct, bg=bg)
        r += 1
    r += 1

    ws2.cell(row=r, column=1,
             value="RULE 4 — TIER 4 FLAG QUALITY").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Metric", "Count", "% of Flagged", "Explanation"], 1):
        _hdr(ws2, r, ci, h, bg=C_BLUE)
    r += 1
    pct_e  = f"{n_r4a_errors/n_r4a_total*100:.1f}%"    if (n_r4a_total and n_r4a_errors is not None)     else "N/A"
    pct_fa = f"{n_r4a_false_alarm/n_r4a_total*100:.1f}%" if (n_r4a_total and n_r4a_false_alarm is not None) else "N/A"
    for label, cnt, pct, expl in [
        ("Total referred (Rule 4)",          n_r4a_total,       "100%",  "Isolated LOW/Mixed not corrected/clustered → Tier 4"),
        ("True errors in referrals",          n_r4a_errors,       pct_e,   "Referred AND actually wrong"),
        ("False referrals (CNN was right)",   n_r4a_false_alarm,  pct_fa,  "Referred BUT CNN was correct"),
    ]:
        bg = C_GB if "True" in label else (C_RB if "False" in label else C_AB)
        _cell(ws2, r, 1, label, bold=True, align="left")
        _cell(ws2, r, 2, cnt if cnt is not None else "N/A", bg=bg)
        _cell(ws2, r, 3, pct, bg=bg)
        _cell(ws2, r, 4, expl, align="left")
        r += 1
    r += 1

    ws2.cell(row=r, column=1,
             value="RULE 3 — TIER 3 CLUSTER FLAG QUALITY").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Metric", "Count", "% of R3 Flagged", "Explanation"], 1):
        _hdr(ws2, r, ci, h, bg=C_BLUE)
    r += 1
    pct_r3e = f"{n_r3_errors/n_r3_flagged*100:.1f}%" if n_r3_flagged > 0 else "N/A"
    for label, cnt, pct, expl in [
        ("Total flagged by Rule 3",   n_r3_flagged,              "100%",   "Cluster flag"),
        ("True errors in cluster",    n_r3_errors,               pct_r3e,  "Flagged AND actually wrong"),
        ("False alarms in cluster",   n_r3_flagged - n_r3_errors if has_gt else None,
         f"{(n_r3_flagged-n_r3_errors)/n_r3_flagged*100:.1f}%" if (has_gt and n_r3_flagged) else "N/A",
         "Flagged but CNN was correct"),
    ]:
        bg = C_GB if "True" in label else (C_AB if "Total" in label else C_GREY)
        _cell(ws2, r, 1, label, bold=True, align="left")
        _cell(ws2, r, 2, cnt if cnt is not None else "N/A", bg=bg)
        _cell(ws2, r, 3, pct, bg=bg)
        _cell(ws2, r, 4, expl, align="left")
        r += 1
    r += 1

    ws2.cell(row=r, column=1,
             value="SILENT FAILURES (firm wrong AND not flagged)").font = Font(bold=True, size=11, color=C_BLUE)
    r += 1
    for ci, h in enumerate(["Metric", "Count", "% of Firm Total", "Explanation"], 1):
        _hdr(ws2, r, ci, h, bg=C_BLUE)
    r += 1
    n_pr_firm_total = deferral_stats.get("n_pr_firm", n_total)
    pct_uf = f"{n_unflagged/n_pr_firm_total*100:.1f}%" if (n_unflagged is not None and n_pr_firm_total) else "N/A"
    _cell(ws2, r, 1, "Silent failures (unflagged firm errors)", bold=True, align="left")
    _cell(ws2, r, 2, n_unflagged if n_unflagged is not None else "N/A", bg=C_RB, ft=C_RT, bold=True)
    _cell(ws2, r, 3, pct_uf, bg=C_RB, ft=C_RT)
    _cell(ws2, r, 4, "Firm wrong AND not flagged. These are the true misses.", align="left")
    r += 2

    if has_gt:
        ws2.cell(row=r, column=1,
                 value="IMPROVEMENT BY IMAGE TYPE  (firm predictions only)").font = Font(
            bold=True, size=11, color=C_BLUE)
        r += 1
        for ci, h in enumerate(["Type","N total","N firm","N deferred",
                                 "ML Acc","PR Acc","Delta",
                                 "Changed","Improved","Worsened"], 1):
            _hdr(ws2, r, ci, h, bg=C_BLUE)
        r += 1
        for tname in ["individual", "sparse_grid", "grid"]:
            s  = type_stats.get(tname, {})
            bg = type_bg_cov[tname]; ft = type_ft_cov[tname]
            _cell(ws2, r, 1, tname,                bg=bg, ft=ft, bold=True, align="left")
            _cell(ws2, r, 2, s.get("n",""),        bg=bg, ft=ft)
            _cell(ws2, r, 3, s.get("n_firm",""),   bg=bg, ft=ft)
            _cell(ws2, r, 4, s.get("n_deferred",""), bg=C_TEAL, ft=C_TEALT)
            _cell(ws2, r, 5, s.get("ml_acc","N/A"), bg=bg, ft=ft)
            _cell(ws2, r, 6, s.get("pr_acc","N/A"), bg=bg, ft=ft)
            dv = s.get("delta","N/A")
            dv_bg = (C_GB if isinstance(dv,str) and dv.startswith("+") else
                     C_RB if isinstance(dv,str) and dv.startswith("-") else bg)
            _cell(ws2, r, 7,  dv,                  bg=dv_bg, bold=True)
            _cell(ws2, r, 8,  s.get("changed",""))
            _cell(ws2, r, 9,  s.get("improved",""), bg=C_GB, ft=C_GT)
            _cell(ws2, r, 10, s.get("worsened",""), bg=C_RB, ft=C_RT)
            r += 1

    _auto_w(ws2)
    ws2.column_dimensions["E"].width = 55

    # ─────────────────────────────────────────────────────────────
    #  Sheet: Triage Tiers
    # ─────────────────────────────────────────────────────────────
    ws_triage = wb.create_sheet("Triage Tiers")
    ws_triage.sheet_view.showGridLines = False
    t_tr = ws_triage.cell(row=1, column=1, value="Triage Tier Summary")
    t_tr.font      = Font(bold=True, size=13, color=C_WHITE)
    t_tr.fill      = _fill(C_NAVY)
    t_tr.alignment = Alignment(horizontal="center", vertical="center")
    ws_triage.merge_cells("A1:G1")
    ws_triage.row_dimensions[1].height = 24

    rtr = 3
    for ci, h in enumerate(["Tier","Label","N patches","% of eligible",
                             "N changed","N flagged"], 1):
        _hdr(ws_triage, rtr, ci, h, bg=C_NAVY, sz=10)
    rtr += 1

    n_elig_triage = max(n_tier1 + n_tier2 + n_tier3 + n_tier4, 1)
    tier_defs = [
        (1, "Confirmed", n_tier1, C_GB,   C_GT),
        (2, "Suggested", n_tier2, C_AB,   C_AT),
        (3, "Attention", n_tier3, C_TEAL, C_TEALT),
        (4, "Refer",     n_tier4, C_RB,   C_RT),
    ]
    for tier_num, tier_lbl, tier_n, bg_t, ft_t in tier_defs:
        pct      = f"{tier_n / n_elig_triage * 100:.1f}%" if n_elig_triage else "0%"
        tp_patch = patch_df[patch_df["triage_tier"] == tier_num] if "triage_tier" in patch_df.columns else pd.DataFrame()
        n_chg_t  = int(tp_patch["changed"].sum())       if len(tp_patch) else 0
        n_flg_t  = int(tp_patch["needs_review"].sum())  if len(tp_patch) else 0
        _cell(ws_triage, rtr, 1, tier_num,  bold=True, bg=bg_t, ft=ft_t)
        _cell(ws_triage, rtr, 2, tier_lbl,  bold=True, bg=bg_t, ft=ft_t, align="left")
        _cell(ws_triage, rtr, 3, tier_n,    bold=True, bg=bg_t, ft=ft_t)
        _cell(ws_triage, rtr, 4, pct,       bg=bg_t, ft=ft_t)
        _cell(ws_triage, rtr, 5, n_chg_t,   bg=C_AB if n_chg_t else None)
        _cell(ws_triage, rtr, 6, n_flg_t,   bg=C_AB if n_flg_t else None)
        rtr += 1

    n_tot_triage = n_tier1 + n_tier2 + n_tier3 + n_tier4
    _cell(ws_triage, rtr, 1, "TOTAL", bold=True, align="left")
    _cell(ws_triage, rtr, 3, n_tot_triage, bold=True)
    _cell(ws_triage, rtr, 4, "100%", bold=True)
    rtr += 2

    ws_triage.cell(row=rtr, column=1,
                   value="RULE BREAKDOWN BY TIER").font = Font(bold=True, size=10, color=C_BLUE)
    rtr += 1
    for ci, h in enumerate(["Rule","Total fired","In Tier 1","In Tier 2",
                             "In Tier 3","In Tier 4","Action type"], 1):
        _hdr(ws_triage, rtr, ci, h, bg=C_BLUE)
    rtr += 1
    rule_action = {
        "rule1a":         "Prediction changed",
        "rule1b":         "Prediction changed",
        "rule2_neighbor": "Prediction changed",
        "rule3_cluster_flag": "Flag only",
        "rule4_refer":    "Flag only",
    }
    if "rule_applied" in patch_df.columns and "triage_tier" in patch_df.columns:
        for rule_name, action in rule_action.items():
            sub_rule  = patch_df[patch_df["rule_applied"].str.contains(
                rule_name.split("_")[0], na=False)]
            cnt_total = len(sub_rule)
            if cnt_total == 0:
                continue
            cnt_t1 = int((sub_rule["triage_tier"] == 1).sum())
            cnt_t2 = int((sub_rule["triage_tier"] == 2).sum())
            cnt_t3 = int((sub_rule["triage_tier"] == 3).sum())
            cnt_t4 = int((sub_rule["triage_tier"] == 4).sum())
            bg_act = C_AB if action == "Prediction changed" else C_GREY
            _cell(ws_triage, rtr, 1, rule_name, bold=True, align="left")
            _cell(ws_triage, rtr, 2, cnt_total, bg=bg_act)
            _cell(ws_triage, rtr, 3, cnt_t1,    bg=C_GB  if cnt_t1 else None)
            _cell(ws_triage, rtr, 4, cnt_t2,    bg=C_AB  if cnt_t2 else None)
            _cell(ws_triage, rtr, 5, cnt_t3,    bg=C_TEAL if cnt_t3 else None, ft=C_TEALT if cnt_t3 else C_BLACK)
            _cell(ws_triage, rtr, 6, cnt_t4,    bg=C_RB  if cnt_t4 else None, ft=C_RT    if cnt_t4 else C_BLACK)
            _cell(ws_triage, rtr, 7, action,    align="left")
            rtr += 1

    if has_gt and "outcome" in patch_df.columns and "triage_tier" in patch_df.columns:
        rtr += 1
        ws_triage.cell(row=rtr, column=1,
            value="OUTCOME BY TIER (requires GT)").font = Font(bold=True, size=10, color=C_BLUE)
        rtr += 1
        for ci, h in enumerate(["Tier","Improved","Error","No change","Unknown (no GT)"], 1):
            _hdr(ws_triage, rtr, ci, h, bg=C_BLUE)
        rtr += 1
        for tier_num, tier_lbl, _, bg_t, ft_t in tier_defs:
            sub_t = patch_df[patch_df["triage_tier"] == tier_num]
            n_imp = int((sub_t["outcome"] == "improved").sum())
            n_wor = int((sub_t["outcome"] == "error").sum())
            n_noc = int((sub_t["outcome"] == "no_change").sum())
            n_unk = int((sub_t["outcome"] == "unknown").sum())
            _cell(ws_triage, rtr, 1, f"Tier {tier_num} {tier_lbl}",
                  bold=True, align="left", bg=bg_t, ft=ft_t)
            _cell(ws_triage, rtr, 2, n_imp, bg=C_GB if n_imp else None, ft=C_GT if n_imp else C_BLACK)
            _cell(ws_triage, rtr, 3, n_wor, bg=C_RB if n_wor else None, ft=C_RT if n_wor else C_BLACK)
            _cell(ws_triage, rtr, 4, n_noc)
            _cell(ws_triage, rtr, 5, n_unk, bg=C_GREY)
            rtr += 1

    _auto_w(ws_triage)

    # ─────────────────────────────────────────────────────────────
    #  Sheets: Tier 1-4 detail
    # ─────────────────────────────────────────────────────────────
    TIER_CONFIG = [
        (1, "Tier 1 — Confirmed", C_GB,   C_GT,
         "HIGH-confidence firm, not corrected, not flagged. Spot-check depth only."),
        (2, "Tier 2 — Suggested", C_AB,   C_AT,
         "Rule-corrected OR rule-supported. Suggested label available."),
        (3, "Tier 3 — Attention", C_TEAL, C_TEALT,
         "Rule 3 cluster flag. Spatially coherent uncertainty/minority zone."),
        (4, "Tier 4 — Refer",     C_RB,   C_RT,
         "Rule 4 isolated LOW/Mixed flag. Expert decides independently."),
    ]
    TIER_COLS = [
        "image_name","slide_id","row","col",
        "gt_label","ml_predicted","ml_conf","ml_conf_tier",
        "final_predicted","suggested_label","pr_zone",
        "changed","rule_applied","needs_review",
        "ml_correct","pr_correct","outcome",
        "triage_tier","triage_label",
    ]
    for tier_num, sheet_title, bg_hdr, ft_hdr, tier_desc in TIER_CONFIG:
        ws_t = wb.create_sheet(sheet_title[:31])
        ws_t.sheet_view.showGridLines = False
        n_cols    = len([c for c in TIER_COLS if c in patch_df.columns])
        title_c   = ws_t.cell(row=1, column=1, value=sheet_title)
        title_c.font      = Font(bold=True, size=13, color=C_WHITE)
        title_c.fill      = _fill(C_NAVY)
        title_c.alignment = Alignment(horizontal="center", vertical="center")
        ws_t.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        ws_t.row_dimensions[1].height = 22
        desc_c = ws_t.cell(row=2, column=1, value=tier_desc)
        desc_c.font      = Font(italic=True, size=9, color="595959")
        desc_c.alignment = Alignment(wrap_text=True, horizontal="left")
        ws_t.merge_cells(f"A2:{get_column_letter(n_cols)}2")
        ws_t.row_dimensions[2].height = 28

        tier_df = patch_df[patch_df["triage_tier"] == tier_num].copy()
        if len(tier_df) == 0:
            ws_t.cell(row=3, column=1, value="No patches in this tier.")
            continue

        n_firm_t  = int((tier_df["pr_zone"] == "firm").sum())
        n_defer_t = int((tier_df["pr_zone"] == "deferred").sum())
        n_corr_t  = int(tier_df["changed"].sum())
        n_flag_t  = int(tier_df["needs_review"].sum())
        n_wrong_t = int((tier_df["pr_correct"] == False).sum()) if has_gt else None
        n_ok_t    = int((tier_df["pr_correct"] == True).sum())  if has_gt else None

        parts = [f"N={len(tier_df)}", f"firm={n_firm_t}", f"deferred={n_defer_t}",
                 f"corrected={n_corr_t}", f"flagged={n_flag_t}"]
        if has_gt:
            parts += [f"correct={n_ok_t}", f"wrong={n_wrong_t}",
                      f"problems={n_wrong_t + n_defer_t}"]
        sum_c = ws_t.cell(row=3, column=1, value="  |  ".join(parts))
        sum_c.font      = Font(bold=True, size=9, color=ft_hdr)
        sum_c.fill      = _fill(bg_hdr)
        sum_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_t.merge_cells(f"A3:{get_column_letter(n_cols)}3")
        ws_t.row_dimensions[3].height = 20

        avail_cols = [c for c in TIER_COLS if c in tier_df.columns]
        for ci, col in enumerate(avail_cols, 1):
            _hdr(ws_t, 4, ci, col, bg=bg_hdr, ft=C_WHITE)

        for ri, (_, row) in enumerate(tier_df.iterrows(), 5):
            outcome = str(row.get("outcome", ""))
            pr_zone = str(row.get("pr_zone", ""))
            pr_ok   = row.get("pr_correct")
            changed = bool(row.get("changed", False))
            if outcome == "improved":    row_bg = C_GB
            elif outcome == "worsened":  row_bg = C_RB
            elif pr_zone == "deferred":  row_bg = C_TEAL
            elif pr_ok is False:         row_bg = C_RB
            elif changed:                row_bg = C_AB
            elif ri % 2 == 0:            row_bg = C_GREY
            else:                        row_bg = None

            for ci, col in enumerate(avail_cols, 1):
                v   = row.get(col)
                bg  = row_bg;  ft = "000000";  bld = False
                fmt = "0.0000" if col in ("ml_conf","ml_prob_G","ml_prob_Gplus",
                                           "majority_ratio","slide_ratio") else None
                if col == "outcome":
                    if v == "improved":            bg, ft, bld = C_GB, C_GT, True
                    elif v == "worsened":          bg, ft, bld = C_RB, C_RT, True
                if col == "pr_correct":
                    if v is True:  bg = C_GB
                    elif v is False: bg = C_RB
                if col == "ml_correct":
                    if v is True:  bg = C_GB
                    elif v is False: bg = C_RB
                if col == "pr_zone" and v == "deferred": bg, ft = C_TEAL, C_TEALT
                if col == "changed" and v is True:       bg, bld = C_AB, True
                if col == "needs_review" and v is True:  bg = C_AB
                _cell(ws_t, ri, ci, v, bg=bg, ft=ft, bold=bld, fmt=fmt,
                      align="left" if col in ("image_name","slide_id","rule_applied",
                                               "outcome","gt_label","ml_predicted",
                                               "final_predicted","triage_label") else "center")

        _auto_w(ws_t)
        ws_t.freeze_panes = "A5"
        ws_t.column_dimensions["A"].width = 28

    # ─────────────────────────────────────────────────────────────
    #  Data sheets
    # ─────────────────────────────────────────────────────────────
    slide_df_display = slide_df.drop(
        columns=[c for c in ("ml_correct","pr_correct") if c in slide_df.columns],
        errors="ignore"
    )
    for title, df, kwargs in [
        ("Patch Results",    patch_df,
         dict(outcome_col="outcome", changed_col="changed", correct_col="pr_correct")),
        ("Changed Patches",  changed_df,
         dict(outcome_col="outcome", changed_col="changed", correct_col="pr_correct")),
        ("Review Flags",     review_df,
         dict(correct_col="ml_correct")),
        ("Errors Remaining",
         errors_df if has_gt else pd.DataFrame(),
         dict(outcome_col="outcome", changed_col="changed")),
    ]:
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        if len(df) == 0:
            ws.cell(row=1, column=1, value=f"No data for: {title}")
        else:
            _write_df(ws, df, **kwargs)
        _auto_w(ws)
        if len(df) > 0:
            ws.freeze_panes = "C2"

    # ─────────────────────────────────────────────────────────────
    #  Sheet: Workload & Triage  (clinical framing summary)
    # ─────────────────────────────────────────────────────────────
    ws_wl2 = wb.create_sheet("Workload & Triage")
    ws_wl2.sheet_view.showGridLines = False

    wl2_title = ws_wl2.cell(row=1, column=1,
        value="WORKLOAD REDUCTION & TRIAGE QUALITY  (primary clinical claims)")
    wl2_title.font      = Font(bold=True, size=13, color=C_WHITE)
    wl2_title.fill      = _fill(C_NAVY)
    wl2_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_wl2.merge_cells("A1:F1")
    ws_wl2.row_dimensions[1].height = 24

    wl2_note = ws_wl2.cell(row=2, column=1, value=(
        "Primary clinical question: does the spatial rule engine reduce the expert's "
        "review burden without increasing diagnostic risk? "
        "NNR = Number Needed to Review: patches an expert must examine to find one error. "
        "Lower NNR = errors are more concentrated = expert time is better directed."
    ))
    wl2_note.font      = Font(italic=True, size=9, color="595959")
    wl2_note.alignment = Alignment(wrap_text=True, horizontal="left")
    ws_wl2.merge_cells("A2:F2")
    ws_wl2.row_dimensions[2].height = 42

    wm2 = workload_metrics or {}
    rw  = 4

    def _wl2_section(ws, r, title):
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=11, color=C_BLUE)
        ws.merge_cells(f"A{r}:F{r}")
        return r + 1

    def _wl2_hdr(ws, r, cols):
        for ci, h in enumerate(cols, 1):
            _hdr(ws, r, ci, h, bg=C_BLUE)
        return r + 1

    def _wl2_row(ws, r, label, val_ml, val_pr, delta, interpretation,
                 bg_pr=None, bold_pr=False):
        _cell(ws, r, 1, label, bold=True, align="left")
        _cell(ws, r, 2, val_ml, bg=C_AB)
        _cell(ws, r, 3, val_pr,
              bg=bg_pr or C_TEAL,
              ft=C_TEALT if bg_pr is None else C_BLACK,
              bold=bold_pr)
        _cell(ws, r, 4, delta)
        _cell(ws, r, 5, interpretation, align="left")
        return r + 1

    if wm2:
        # ── 1. Review burden reduction ───────────────────────────
        rw = _wl2_section(ws_wl2, rw,
            "1. REVIEW BURDEN REDUCTION")
        rw = _wl2_hdr(ws_wl2, rw,
            ["Metric", "CNN-only", "Rule engine", "Change", "Interpretation"])

        rw = _wl2_row(ws_wl2, rw,
            "Total patches requiring deep review",
            wm2.get("n_total", 0),
            wm2.get("n_tier34", 0),
            f"-{wm2.get('n_total',0) - wm2.get('n_tier34',0)} "
            f"({100 - wm2.get('deep_review_pct',0):.1f}% reduction)",
            "CNN-only: all patches need deep review. "
            "Rule engine: only Tier 3+4 need deep review; Tier 1+2 use reduced/suggested depth.",
            bg_pr=C_GB, bold_pr=True)

        rw = _wl2_row(ws_wl2, rw,
            "Deep-review burden %",
            "100%",
            f"{wm2.get('deep_review_pct', 0):.1f}%",
            f"-{100 - wm2.get('deep_review_pct', 0):.1f}%",
            "% of patches requiring full deep expert attention (Tier 3+4)",
            bg_pr=C_GB if (100 - wm2.get("deep_review_pct", 0)) > 30 else None)

        rw = _wl2_row(ws_wl2, rw,
            "Tier 1 (spot-check) patches",
            "0 (no structured triage)",
            wm2.get("n_tier1", 0),
            f"+{wm2.get('n_tier1', 0)}",
            "Confirmed HIGH-conf, unflagged — reduced inspection depth, not omission",
            bg_pr=C_GB, bold_pr=True)

        rw = _wl2_row(ws_wl2, rw,
            "Tier 2 (suggested label) patches",
            "—", wm2.get("n_tier2", 0), "—",
            "Rule engine provides a suggested label with spatial rationale")

        rw = _wl2_row(ws_wl2, rw,
            "Tier 3 (cluster attention) patches",
            "—", wm2.get("n_tier3", 0), "—",
            "Rule 3 cluster flag — collective expert confirmation required")

        rw = _wl2_row(ws_wl2, rw,
            "Tier 4 (isolated refer) patches",
            "—", wm2.get("n_tier4", 0), "—",
            "Rule 4 isolated uncertainty — full independent human decision",
            bg_pr=C_RB)
        rw += 1

        # ── 2. Spot-check safety (density ratio) ─────────────────
        rw = _wl2_section(ws_wl2, rw,
            "2. SPOT-CHECK QUALITY  (is Tier 1 error-sparse vs Tier 3+4?)")
        rw = _wl2_hdr(ws_wl2, rw,
            ["Metric", "Tier 1", "Tier 3+4", "Assessment", "Interpretation"])

        t1_dens   = wm2.get("t1_error_density", 0)
        t34_dens  = wm2.get("t34_error_density", 0)
        base_dens = wm2.get("baseline_density", 0)
        dr2       = wm2.get("density_ratio")
        dr2_inf   = wm2.get("density_ratio_infinite", False)
        tq2       = wm2.get("triage_quality", "unknown")
        enrich2   = wm2.get("t34_enrichment")

        dr2_str = "∞ (Tier 1 zero errors)" if dr2_inf else (f"{dr2:.2f}×" if dr2 else "N/A")
        tq2_bg  = C_GB if tq2 == "strong" else (C_AB if tq2 == "adequate" else C_RB)
        tq2_ft  = C_GT if tq2 == "strong" else (C_AT if tq2 == "adequate" else C_RT)

        _cell(ws_wl2, rw, 1, "Error density (errors / patches)", bold=True, align="left")
        _cell(ws_wl2, rw, 2, f"{t1_dens*100:.2f}%",
              bg=C_GB if t1_dens < base_dens else C_RB)
        _cell(ws_wl2, rw, 3, f"{t34_dens*100:.2f}%", bg=C_AB)
        _cell(ws_wl2, rw, 4, f"Baseline: {base_dens*100:.2f}%")
        _cell(ws_wl2, rw, 5,
              "Tier 1 should be well below baseline; Tier 3+4 should be above.",
              align="left")
        rw += 1

        _cell(ws_wl2, rw, 1, "Density ratio  (Tier3+4 ÷ Tier1)", bold=True, align="left")
        _cell(ws_wl2, rw, 2, dr2_str, bg=tq2_bg, ft=tq2_ft, bold=True)
        _cell(ws_wl2, rw, 3, tq2.upper(), bg=tq2_bg, ft=tq2_ft, bold=True)
        _cell(ws_wl2, rw, 4, "≥3× = strong")
        _cell(ws_wl2, rw, 5,
              "∞ = Tier 1 perfectly clean. ≥3× = strong triage. <1× = failed.",
              align="left")
        rw += 1

        enrich2_str = "∞" if wm2.get("t34_enrichment_infinite") else \
                      (f"{enrich2:.2f}×" if enrich2 else "N/A")
        enrich2_bg  = C_GB if (enrich2 and enrich2 >= 1.0) or \
                      wm2.get("t34_enrichment_infinite") else C_RB
        _cell(ws_wl2, rw, 1, "Tier 3+4 enrichment vs CNN uniform baseline",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, "1.00× (uniform)")
        _cell(ws_wl2, rw, 3, enrich2_str, bg=enrich2_bg, bold=True)
        _cell(ws_wl2, rw, 4, "≥1×")
        _cell(ws_wl2, rw, 5,
              ">1× = directing experts to Tier 3+4 is more efficient than no triage.",
              align="left")
        rw += 1

        t1_err2     = wm2.get("t1_errors", 0)
        t1_err_rate = wm2.get("t1_error_rate_pct", 0)
        t1_note_bg  = C_GB if t1_err2 == 0 else C_AB
        _cell(ws_wl2, rw, 1, "Tier 1 errors (firm wrong in spot-check pool)",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, t1_err2, bg=t1_note_bg)
        _cell(ws_wl2, rw, 3, f"{t1_err_rate:.2f}%", bg=t1_note_bg)
        _cell(ws_wl2, rw, 4, "Low is good")
        _cell(ws_wl2, rw, 5,
              "0 = best. Low count acceptable if density_ratio is strong.",
              align="left")
        rw += 2

        # ── 3. NNR ───────────────────────────────────────────────
        rw = _wl2_section(ws_wl2, rw,
            "3. NUMBER NEEDED TO REVIEW (NNR)  — clinical efficiency metric")
        rw = _wl2_hdr(ws_wl2, rw,
            ["Metric", "CNN-only", "Rule engine (Tier 3+4)",
             "Improvement", "Interpretation"])

        nnr_ml2  = wm2.get("nnr_ml");  nnr_pr2  = wm2.get("nnr_pr")
        nnr_imp2 = wm2.get("nnr_improvement")
        nnr_bg2  = C_GB if (nnr_ml2 and nnr_pr2 and nnr_pr2 < nnr_ml2) else C_RB

        rw = _wl2_row(ws_wl2, rw,
            "NNR (patches per problem found)",
            round(nnr_ml2, 2) if nnr_ml2 else "N/A",
            round(nnr_pr2, 2) if nnr_pr2 else "N/A",
            f"×{nnr_imp2:.1f} more efficient" if nnr_imp2 else "—",
            "Lower NNR in Tier 3+4 = problems more concentrated = expert finds them faster",
            bg_pr=nnr_bg2, bold_pr=True)

        rw = _wl2_row(ws_wl2, rw,
            "Problems in review pool (errors + deferred)",
            wm2.get("nnr_ml_problems", 0),
            wm2.get("nnr_pr_problems", 0),
            f"{wm2.get('nnr_pr_problems',0) - wm2.get('nnr_ml_problems',0):+d}",
            "Errors + unresolved deferrals the expert must handle in each pool")
        rw += 1

        # ── 4. Problem concentration ─────────────────────────────
        rw = _wl2_section(ws_wl2, rw,
            "4. PROBLEM CONCENTRATION  (are problems pushed into the review zone?)")
        rw = _wl2_hdr(ws_wl2, rw,
            ["Metric", "Count", "%", "Target", "Interpretation"])

        all_prob2   = wm2.get("all_problems", 0)
        t34_prob2   = wm2.get("t34_problems", 0)
        t1_prob2    = wm2.get("t1_problems", 0)
        prob_c_pct2 = wm2.get("problem_concentration_t34", 0)
        prob_t1_pct2= wm2.get("problem_in_t1_pct", 0)
        n_def_t12   = wm2.get("n_pr_deferred_t1", 0)
        n_def_t342  = wm2.get("n_pr_deferred_t34", 0)

        prob_bg2 = C_GB if prob_c_pct2 >= 95 else (C_AB if prob_c_pct2 >= 85 else C_RB)
        t1_bg2   = C_GB if t1_prob2 == 0 else (C_AB if prob_t1_pct2 <= 5 else C_RB)

        _cell(ws_wl2, rw, 1, "All problems — firm wrong + deferred (total)",
              bold=True, align="left", bg=C_PURP, ft=C_PURPT)
        _cell(ws_wl2, rw, 2, all_prob2, bg=C_PURP, ft=C_PURPT, bold=True)
        _cell(ws_wl2, rw, 3, "100%", bg=C_PURP)
        _cell(ws_wl2, rw, 4, "—")
        _cell(ws_wl2, rw, 5, "Wrong predictions + unresolved deferrals", align="left")
        rw += 1

        _cell(ws_wl2, rw, 1, "Problems in Tier 3+4 (deep-review zone)",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, t34_prob2, bg=prob_bg2, bold=True)
        _cell(ws_wl2, rw, 3, f"{prob_c_pct2:.1f}%", bg=prob_bg2, bold=True)
        _cell(ws_wl2, rw, 4, "≥95%")
        _cell(ws_wl2, rw, 5,
              f"Includes {n_def_t342} deferred patches. Expert attention directed here.",
              align="left")
        rw += 1

        _cell(ws_wl2, rw, 1, "Problems in Tier 1 (spot-check zone)",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, t1_prob2, bg=t1_bg2, bold=True)
        _cell(ws_wl2, rw, 3, f"{prob_t1_pct2:.1f}%", bg=t1_bg2, bold=True)
        _cell(ws_wl2, rw, 4, "≤5%")
        _cell(ws_wl2, rw, 5,
              f"Includes {n_def_t12} deferred in Tier 1. "
              "High density ratio mitigates any residual risk.",
              align="left")
        rw += 2

        # ── 5. Flag quality ──────────────────────────────────────
        rw = _wl2_section(ws_wl2, rw,
            "5. FLAG QUALITY  (when the system says 'look here', is it right?)")
        rw = _wl2_hdr(ws_wl2, rw,
            ["Flag source", "Patches flagged", "True errors",
             "True-error rate", "Interpretation"])

        r3_n2   = wm2.get("n_r3_flagged", 0)
        r3_err2 = wm2.get("n_r3_errors", 0)
        r3_ter2 = wm2.get("r3_true_error_rate", 0)
        r3_bg2  = C_GB if r3_ter2 >= 70 else (C_AB if r3_ter2 >= 40 else C_RB)
        _cell(ws_wl2, rw, 1, "Rule 3 — Cluster flag", bold=True, align="left")
        _cell(ws_wl2, rw, 2, r3_n2,   bg=C_AB)
        _cell(ws_wl2, rw, 3, r3_err2, bg=r3_bg2, bold=True)
        _cell(ws_wl2, rw, 4, f"{r3_ter2:.1f}%", bg=r3_bg2, bold=True)
        _cell(ws_wl2, rw, 5,
              "Spatial cluster of uncertain patches. ≥70% = clinically strong signal.",
              align="left")
        rw += 1

        ov_n2   = wm2.get("n_flagged_total", 0)
        ov_err2 = wm2.get("n_flagged_errors", 0)
        ov_pr2  = wm2.get("flag_precision_all", 0)
        ov_bg2  = C_GB if ov_pr2 >= 40 else (C_AB if ov_pr2 >= 20 else C_RB)
        _cell(ws_wl2, rw, 1, "All flags combined (Rule 3 + Rule 4)",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, ov_n2,   bg=C_AB)
        _cell(ws_wl2, rw, 3, ov_err2, bg=ov_bg2)
        _cell(ws_wl2, rw, 4, f"{ov_pr2:.1f}%", bg=ov_bg2)
        _cell(ws_wl2, rw, 5,
              "Rule 4 is intentionally broad (sensitive safety net).",
              align="left")
        rw += 2

        # ── 6. Tier 2 correction quality ─────────────────────────
        rw = _wl2_section(ws_wl2, rw,
            "6. TIER 2 CORRECTION QUALITY  (when spatial rule changes a prediction, is it right?)")
        rw = _wl2_hdr(ws_wl2, rw,
            ["Metric", "Count", "%", "Target", "Interpretation"])

        t2_imp2  = wm2.get("n_t2_improved", 0)
        t2_wor2  = wm2.get("n_t2_worsened", 0)
        t2_prec2 = wm2.get("t2_correction_precision", 0)
        t2_bg2   = C_GB if t2_prec2 >= 85 else (C_AB if t2_prec2 >= 70 else C_RB)
        _cell(ws_wl2, rw, 1, "Tier 2 corrections: improved",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, t2_imp2, bg=C_GB, ft=C_GT, bold=True)
        _cell(ws_wl2, rw, 3, f"{t2_prec2:.1f}%", bg=t2_bg2, bold=True)
        _cell(ws_wl2, rw, 4, "≥85%")
        _cell(ws_wl2, rw, 5,
              "% of rule-corrections that fixed a CNN error", align="left")
        rw += 1

        _cell(ws_wl2, rw, 1, "Tier 2 corrections: error (introduced wrong label)",
              bold=True, align="left")
        _cell(ws_wl2, rw, 2, t2_wor2,
              bg=C_RB if t2_wor2 else None,
              ft=C_RT if t2_wor2 else C_BLACK, bold=bool(t2_wor2))
        _cell(ws_wl2, rw, 3,
              f"{100 - t2_prec2:.1f}%" if (t2_imp2 + t2_wor2) > 0 else "0%",
              bg=C_RB if t2_wor2 else None)
        _cell(ws_wl2, rw, 4, "≤15%")
        _cell(ws_wl2, rw, 5,
              "% of rule-corrections that broke a correct CNN prediction",
              align="left")
        rw += 1

    else:
        ws_wl2.cell(row=4, column=1,
            value="No GT labels — workload metrics require ground truth.")

    _auto_w(ws_wl2)
    ws_wl2.column_dimensions["A"].width = 44
    ws_wl2.column_dimensions["E"].width = 52

    out = os.path.join(output_dir, "evaluation_results.xlsx")
    wb.save(out)
    print(f"  [Excel] Saved: {out}")


# ─────────────────────────────────────────────────────────────────
#  Console report
# ─────────────────────────────────────────────────────────────────

def _print_report(
    pm_ml, pm_pr, sm_ml, sm_pr,
    patch_df, changed_df, slide_df,
    n_individual, n_sparse, n_grid, n_eligible,
    n_r4_total, n_r4_errors, n_r4_false_alarm, n_unflagged,
    has_gt, output_dir,
    mixed_stats=None, deferral_stats=None,
    fair_comparisons=None, workload_metrics=None,
):
    W = 70; L = []
    def h(c="="): L.append(c * W)
    def ln(t=""):  L.append(t)

    h(); ln("SPATIAL RULE ENGINE — EVALUATION REPORT")
    ln(f"  Total patches : {len(patch_df)}")
    ln(f"  Total slides  : {len(slide_df)}")
    ln(f"  Has GT labels : {has_gt}")
    h()

    ln(f"\nIMAGE COVERAGE")
    ln(f"  Individual (rules cannot fire) : {n_individual}")
    ln(f"  Sparse grid (too few patches)  : {n_sparse}")
    ln(f"  Grid patches (rule-eligible)   : {n_grid}")

    n_ch = int(patch_df["changed"].sum())
    n_rv = int(patch_df["needs_review"].sum())
    ln(f"\nRULE ENGINE ACTIONS (on {n_grid} eligible grid patches)")
    ln(f"  Changed (Rule 1/2) : {n_ch}")
    ln(f"  Flagged (Rule 4)   : {n_rv}")
    if n_ch > 0:
        for rule, cnt in patch_df[patch_df["changed"]]["rule_applied"].value_counts().items():
            ln(f"    {rule}: {cnt}")

    if has_gt and n_ch > 0:
        n_imp = int((patch_df[patch_df["changed"]]["outcome"] == "improved").sum())
        n_wor = int((patch_df[patch_df["changed"]]["outcome"] == "error").sum())
        ln(f"\n  Of {n_ch} changed: {n_imp} improved, {n_wor} worsened")

    if has_gt and deferral_stats:
        ds = deferral_stats
        ln(f"\nTABLE B — DEFERRAL & COVERAGE")
        ln(f"  {'Metric':<30} {'ML only':>12} {'Rule engine':>12}")
        ln("  " + "-" * 56)
        ln(f"  {'Total G/Gplus GT patches':<30} {ds.get('n_total',0):>12}")
        ln(f"  {'Firm predictions':<30} {ds.get('n_ml_firm',0):>12} {ds.get('n_pr_firm',0):>12}")
        ln(f"  {'Coverage %':<30} "
           f"{ds.get('ml_coverage',0)*100:>11.1f}% "
           f"{ds.get('pr_coverage',0)*100:>11.1f}%")
        ln(f"  {'Deferred (Mixed output)':<30} "
           f"{ds.get('n_ml_deferred',0):>12} "
           f"{ds.get('n_pr_deferred',0):>12}")
        ln(f"  {'Deferral rate %':<30} "
           f"{ds.get('ml_deferral_rate',0)*100:>11.1f}% "
           f"{ds.get('pr_deferral_rate',0)*100:>11.1f}%")

    fc = fair_comparisons or {}
    if has_gt and fc:
        ln(f"\nTABLE C — FAIR COMPARISON")
        sa = fc.get("scenario_A", {})
        if sa:
            ln(f"\n  Scenario A — Same N={sa.get('n',0)} (ML-firm patches only):")
            ln(f"  {'Metric':<14} {'ML':>16} {'PR':>14} {'Delta':>10}")
            ln("  " + "-" * 56)
            for key, label in [("mcc","MCC"),("bal_acc","Bal.Acc"),("recall","Recall"),("f1","F1")]:
                ml_v = sa.get("ml",{}).get(key, 0)
                pr_v = sa.get("pr",{}).get(key, 0)
                ln(f"  {label:<14} {ml_v:>16.4f} {pr_v:>14.4f} {pr_v-ml_v:>+10.4f}")
        sb = fc.get("scenario_B", {})
        if sb:
            ln(f"\n  Scenario B — All {sb.get('n',0)} patches, Mixed=wrong:")
            for key, label in [("mcc","MCC"),("bal_acc","Bal.Acc"),("recall","Recall"),("f1","F1")]:
                ml_v = sb.get("ml",{}).get(key, 0)
                pr_v = sb.get("pr",{}).get(key, 0)
                ln(f"  {label:<14} {ml_v:>16.4f} {pr_v:>14.4f} {pr_v-ml_v:>+10.4f}")
        sc = fc.get("scenario_C", {})
        if sc and sc.get("n", 0) > 0:
            acc = sc.get("accuracy", 0)
            ln(f"\n  Scenario C — Extra {sc.get('n',0)} patches PR committed:")
            ln(f"    Correct: {sc.get('n_correct',0)}  Wrong: {sc.get('n_wrong',0)}  "
               f"Accuracy: {acc*100:.1f}%")

    if has_gt and pm_ml:
        ds = deferral_stats or {}
        n_ml_f = ds.get("n_ml_firm", pm_ml.get("n", 0))
        n_pr_f = ds.get("n_pr_firm", pm_pr.get("n", 0))
        ln(f"\nTABLE A — DIAGNOSTIC PERFORMANCE (ML N={n_ml_f}, PR N={n_pr_f})")
        ln(f"  {'Metric':<16} {'ML-only':>10} {'Rule engine':>10} {'Delta':>10}")
        ln("  " + "-" * 50)
        for key, label in [("mcc","MCC"),("bal_acc","Bal.Acc"),("recall","Recall"),
                            ("precision","Precision"),("f1","F1"),("f2","F2"),("acc","Accuracy")]:
            ml_v = pm_ml.get(key, 0); pr_v = pm_pr.get(key, 0)
            ln(f"  {label:<16} {ml_v:>10.4f} {pr_v:>10.4f} {pr_v-ml_v:>+10.4f}")

    if has_gt and pm_ml:
        ln(f"\nCONFUSION MATRIX  (firm predictions only)")
        for lbl, pm in [("ML only", pm_ml), ("Rule engine", pm_pr)]:
            n_f = int(pm.get("tp",0)+pm.get("tn",0)+pm.get("fp",0)+pm.get("fn",0))
            ln(f"  {lbl:<16} TP={pm.get('tp',0)} TN={pm.get('tn',0)} "
               f"FP={pm.get('fp',0)} FN={pm.get('fn',0)} N={n_f}")

    if has_gt:
        ln(f"\nRULE 4 FLAG QUALITY")
        ln(f"  Total flagged  : {n_r4_total}")
        ln(f"  True errors    : {n_r4_errors}")
        ln(f"  False alarms   : {n_r4_false_alarm}")
        ln(f"\nSILENT FAILURES: {n_unflagged}")

    if sm_ml:
        ln(f"\nSLIDE-LEVEL METRICS  (firm diagnoses only)")
        for key, label in [("mcc","MCC"),("bal_acc","Bal.Acc"),("f1","F1"),("acc","Accuracy")]:
            ml_v = sm_ml.get(key, 0); pr_v = sm_pr.get(key, 0)
            ln(f"  {label:<12} {ml_v:>10.4f} {pr_v:>10.4f} {pr_v-ml_v:>+10.4f}")

    if has_gt and workload_metrics:
        wm = workload_metrics
        h(); ln("WORKLOAD REDUCTION & TRIAGE QUALITY"); h("-")
        ln(f"\n1. TIER SPLIT")
        ln(f"   Tier 1 (spot-check) : {wm.get('n_tier1',0)}  ({wm.get('spot_check_pct',0):.1f}%)")
        ln(f"   Tier 2 (suggested)  : {wm.get('n_tier2',0)}")
        ln(f"   Tier 3+4 (deep)     : {wm.get('n_tier34',0)} ({wm.get('deep_review_pct',0):.1f}%)")
        dr     = wm.get("density_ratio"); dr_inf = wm.get("density_ratio_infinite", False)
        tq     = wm.get("triage_quality", "unknown")
        dr_str = "inf" if dr_inf else (f"{dr:.2f}x" if dr else "N/A")
        ln(f"\n2. ERROR DENSITY")
        ln(f"   Baseline : {wm.get('baseline_density',0)*100:.2f}%")
        ln(f"   Tier 1   : {wm.get('t1_error_density',0)*100:.2f}%")
        ln(f"   Tier 3+4 : {wm.get('t34_error_density',0)*100:.2f}%")
        ln(f"   Ratio    : {dr_str}  [{tq.upper()}]")
        ln(f"\n3. NNR")
        nnr_ml_v = wm.get("nnr_ml"); nnr_pr_v = wm.get("nnr_pr")
        ln(f"   CNN-only : {nnr_ml_v:.2f}" if nnr_ml_v else "   CNN-only : N/A")
        ln(f"   Tier 3+4 : {nnr_pr_v:.2f}" if nnr_pr_v else "   Tier 3+4 : N/A")
        imp = wm.get("nnr_improvement")
        if imp:
            ln(f"   → x{imp:.1f} more efficient")
        ln(f"\n4. PROBLEM CONCENTRATION")
        ln(f"   All problems : {wm.get('all_problems',0)}")
        ln(f"   In Tier 3+4  : {wm.get('t34_problems',0)} ({wm.get('problem_concentration_t34',0):.1f}%)")
        ln(f"   In Tier 1    : {wm.get('t1_problems',0)} ({wm.get('problem_in_t1_pct',0):.1f}%)")
        ln(f"\n5. TIER 2 CORRECTION")
        ln(f"   Improved : {wm.get('n_t2_improved',0)}  Error : {wm.get('n_t2_worsened',0)}")
        ln(f"   Precision: {wm.get('t2_correction_precision',0):.1f}%")

    h()
    ln("  evaluation_results.xlsx — full report")
    h()
    text = "\n".join(L)
    print("\n" + text)
    with open(f"{output_dir}/report.txt", "w", encoding="utf-8") as f:
        f.write(text)
